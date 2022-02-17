from typing import MappingView
import pyodbc
import openpyxl

server = "172.30.1.2"
port = 5432
database = "OEE_DB"
username = "sa"
password = "p@ssw0rd"

def SAP_INF_OEE01() :
    path = "D:\\Work\\foster\\oee\\web server V2.1\\web\\dataSAP\\INF_OEE01.xlsx"
 
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        print(cell_obj.value)

        INF_OEE01 = cell_obj.value

        print(INF_OEE01.split("|"))
        
        INF_OEE01_Plant = INF_OEE01.split("|")[0]
        INF_OEE01_WC_Text = INF_OEE01.split("|")[1]
        INF_OEE01_Machine_text = INF_OEE01.split("|")[2]
        INF_OEE01_Machine_ID = INF_OEE01.split("|")[3]
        INF_OEE01_Material = INF_OEE01.split("|")[4]
        INF_OEE01_Description = INF_OEE01.split("|")[5]
        INF_OEE01_PD_Order = INF_OEE01.split("|")[6]
        INF_OEE01_UOM_Qty = INF_OEE01.split("|")[7]
        INF_OEE01_Plan_quantity = INF_OEE01.split("|")[8]
        INF_OEE01_Batch = INF_OEE01.split("|")[9]
        INF_OEE01_Bulk_Code = INF_OEE01.split("|")[10]
        INF_OEE01_Bulk_PD_order1 = INF_OEE01.split("|")[11]
        INF_OEE01_Bulk_PD_order2 = INF_OEE01.split("|")[12]
        INF_OEE01_Bulk_PD_order3 = INF_OEE01.split("|")[13]
        INF_OEE01_Bulk_PD_order4 = INF_OEE01.split("|")[14]
        INF_OEE01_Bulk_PD_order5 = INF_OEE01.split("|")[15]
        INF_OEE01_Bulk_PD_order6 = INF_OEE01.split("|")[16]
        INF_OEE01_Bulk_PD_order7 = INF_OEE01.split("|")[17]
        INF_OEE01_Bulk_PD_order8 = INF_OEE01.split("|")[18]
        INF_OEE01_Bulk_PD_order9 = INF_OEE01.split("|")[19]
        INF_OEE01_Bulk_PD_order10 = INF_OEE01.split("|")[20]
        
        print("INF_OEE01_Plant --> " ,INF_OEE01_Plant )
        print("INF_OEE01_WC_Text --> " ,INF_OEE01_WC_Text )
        print("INF_OEE01_Machine_text --> " ,INF_OEE01_Machine_text )
        print("INF_OEE01_Machine_ID --> " ,INF_OEE01_Machine_ID )
        print("INF_OEE01_Material --> " ,INF_OEE01_Material )
        print("INF_OEE01_Description --> " ,INF_OEE01_Description )
        print("INF_OEE01_PD_Order --> " ,INF_OEE01_PD_Order )
        print("INF_OEE01_UOM_Qty --> " ,INF_OEE01_UOM_Qty )
        print("INF_OEE01_Plan_quantity --> " ,INF_OEE01_Plan_quantity )
        print("INF_OEE01_Batch --> " ,INF_OEE01_Batch )
        print("INF_OEE01_Bulk_Code --> " ,INF_OEE01_Bulk_Code )
        print("INF_OEE01_Bulk_PD_order1 --> " ,INF_OEE01_Bulk_PD_order1 )
        print("INF_OEE01_Bulk_PD_order2 --> " ,INF_OEE01_Bulk_PD_order2 )
        print("INF_OEE01_Bulk_PD_order3 --> " ,INF_OEE01_Bulk_PD_order3 )
        print("INF_OEE01_Bulk_PD_order4 --> " ,INF_OEE01_Bulk_PD_order4 )
        print("INF_OEE01_Bulk_PD_order5 --> " ,INF_OEE01_Bulk_PD_order5 )
        print("INF_OEE01_Bulk_PD_order6 --> " ,INF_OEE01_Bulk_PD_order6 )
        print("INF_OEE01_Bulk_PD_order7 --> " ,INF_OEE01_Bulk_PD_order7 )
        print("INF_OEE01_Bulk_PD_order8 --> " ,INF_OEE01_Bulk_PD_order8 )
        print("INF_OEE01_Bulk_PD_order9 --> " ,INF_OEE01_Bulk_PD_order9 )
        print("INF_OEE01_Bulk_PD_order10 --> " ,INF_OEE01_Bulk_PD_order10 )
        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        INSERT_INF_OEE01 = cnxn.cursor()
        INSERT_INF_OEE01.execute('INSERT INTO OEE_DB.dbo.INF_OEE01 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,UOMQty,PlanQuantity,Batch,BulkCode,BulkPDOrder1,BulkPDOrder2,BulkPDOrder3,BulkPDOrder4,BulkPDOrder5,BulkPDOrder6,BulkPDOrder7,BulkPDOrder8,BulkPDOrder9,BulkPDOrder10) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE01_Plant  , INF_OEE01_WC_Text , INF_OEE01_Machine_text  , INF_OEE01_Machine_ID  , INF_OEE01_Material  , INF_OEE01_Description,INF_OEE01_PD_Order,INF_OEE01_UOM_Qty,INF_OEE01_Plan_quantity,INF_OEE01_Batch,INF_OEE01_Bulk_Code,INF_OEE01_Bulk_PD_order1,INF_OEE01_Bulk_PD_order2,INF_OEE01_Bulk_PD_order3,INF_OEE01_Bulk_PD_order4,INF_OEE01_Bulk_PD_order5,INF_OEE01_Bulk_PD_order6,INF_OEE01_Bulk_PD_order7,INF_OEE01_Bulk_PD_order8,INF_OEE01_Bulk_PD_order9,INF_OEE01_Bulk_PD_order10 ))
        cnxn.commit()
    
def SAP_INF_OEE02() :

    path = "D:\\Work\\foster\\oee\\web server V2.1\\web\\dataSAP\\INF_OEE02.xlsx"
 
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        print(cell_obj.value)

        INF_OEE02 = cell_obj.value
        
    print(INF_OEE02.split("|"))
    
    INF_OEE02_Plant = INF_OEE02.split("|")[0]
    INF_OEE02_WC_Text = INF_OEE02.split("|")[1]
    INF_OEE02_Machine_text = INF_OEE02.split("|")[2]
    INF_OEE02_Machine_ID = INF_OEE02.split("|")[3]
    INF_OEE02_Material = INF_OEE02.split("|")[4]
    INF_OEE02_Description = INF_OEE02.split("|")[5]
    INF_OEE02_PD_Order = INF_OEE02.split("|")[6]
    INF_OEE02_UOM_Qty = INF_OEE02.split("|")[7]
    INF_OEE02_Plan_quantity = INF_OEE02.split("|")[8]
    INF_OEE02_Batch_Estimate = INF_OEE02.split("|")[9]
    INF_OEE02_Down_time_code1  = INF_OEE02.split("|")[10]
    INF_OEE02_Down_time_description1 = INF_OEE02.split("|")[11]
    INF_OEE02_Hrs1 = INF_OEE02.split("|")[12]
    INF_OEE02_Down_time_code2  = INF_OEE02.split("|")[13]
    INF_OEE02_Down_time_description2 = INF_OEE02.split("|")[14]
    INF_OEE02_Hrs2 = INF_OEE02.split("|")[15]
    INF_OEE02_Down_time_code3  = INF_OEE02.split("|")[16]
    INF_OEE02_Down_time_description3 = INF_OEE02.split("|")[17]
    INF_OEE02_Hrs3 = INF_OEE02.split("|")[18]
    INF_OEE02_Down_time_code4  = INF_OEE02.split("|")[19]
    INF_OEE02_Down_time_description4 = INF_OEE02.split("|")[20]
    INF_OEE02_Hrs4 = INF_OEE02.split("|")[21]
    INF_OEE02_Down_time_code5  = INF_OEE02.split("|")[22]
    INF_OEE02_Down_time_description5 = INF_OEE02.split("|")[23]
    INF_OEE02_Hrs5 = INF_OEE02.split("|")[24]
    

    
    print("INF_OEE02_Plant --> " ,INF_OEE02_Plant )
    print("INF_OEE02_WC_Text --> " ,INF_OEE02_WC_Text )
    print("INF_OEE02_Machine_text --> " ,INF_OEE02_Machine_text )
    print("INF_OEE02_Machine_ID --> " ,INF_OEE02_Machine_ID )
    print("INF_OEE02_Material --> " ,INF_OEE02_Material )
    print("INF_OEE02_Description --> " ,INF_OEE02_Description )
    print("INF_OEE02_PD_Order --> " ,INF_OEE02_PD_Order )
    print("INF_OEE02_UOM_Qty --> " ,INF_OEE02_UOM_Qty )
    print("INF_OEE02_Plan_quantity --> " ,INF_OEE02_Plan_quantity )
    print("INF_OEE02_Batch_Estimate --> " ,INF_OEE02_Batch_Estimate )
    print("INF_OEE02_Down_time_code1 --> " ,INF_OEE02_Down_time_code1 )
    print("INF_OEE02_Down_time_description1 --> " ,INF_OEE02_Down_time_description1 )
    print("INF_OEE02_Hrs1 --> " ,INF_OEE02_Hrs1 )
    print("INF_OEE02_Down_time_code2 --> " ,INF_OEE02_Down_time_code2 )
    print("INF_OEE02_Down_time_description2 --> " ,INF_OEE02_Down_time_description2 )
    print("INF_OEE02_Hrs2 --> " ,INF_OEE02_Hrs2 )
    print("INF_OEE02_Down_time_code3 --> " ,INF_OEE02_Down_time_code3 )
    print("INF_OEE02_Down_time_description3 --> " ,INF_OEE02_Down_time_description3 )
    print("INF_OEE02_Hrs3 --> " ,INF_OEE02_Hrs3 )
    print("INF_OEE02_Down_time_code4 --> " ,INF_OEE02_Down_time_code4 )
    print("INF_OEE02_Down_time_description4 --> " ,INF_OEE02_Down_time_description4 )
    print("INF_OEE02_Hrs4 --> " ,INF_OEE02_Hrs4 )
    print("INF_OEE02_Down_time_code5 --> " ,INF_OEE02_Down_time_code5 )
    print("INF_OEE02_Down_time_description5 --> " ,INF_OEE02_Down_time_description5 )
    print("INF_OEE02_Hrs5 --> " ,INF_OEE02_Hrs5 )
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    INSERT_INF_OEE02 = cnxn.cursor()
    INSERT_INF_OEE02.execute('INSERT INTO OEE_DB.dbo.INF_OEE02 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,UOMQty,PlanQuantity,BatchEstimate,DownTimeCode1,DownTimeDescription1,Hrs1,DownTimeCode2,DownTimeDescription2,Hrs2,DownTimeCode3,DownTimeDescription3,Hrs3,DownTimeCode4,DownTimeDescription4,Hrs4,DownTimeCode5,DownTimeDescription5,Hrs5) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE02_Plant  , INF_OEE02_WC_Text , INF_OEE02_Machine_text  , INF_OEE02_Machine_ID  , INF_OEE02_Material  , INF_OEE02_Description,INF_OEE02_PD_Order,INF_OEE02_UOM_Qty,INF_OEE02_Plan_quantity,INF_OEE02_Batch_Estimate,INF_OEE02_Down_time_code1,INF_OEE02_Down_time_description1,INF_OEE02_Hrs1,INF_OEE02_Down_time_code2,INF_OEE02_Down_time_description2,INF_OEE02_Hrs2,INF_OEE02_Down_time_code3,INF_OEE02_Down_time_description3,INF_OEE02_Hrs3,INF_OEE02_Down_time_code4,INF_OEE02_Down_time_description4,INF_OEE02_Hrs4,INF_OEE02_Down_time_code5,INF_OEE02_Down_time_description5,INF_OEE02_Hrs5))
    cnxn.commit()
    
def SAP_INF_OEE03() :

    path = "D:\\Work\\foster\\oee\\web server V2.1\\web\\dataSAP\\INF_OEE03.xlsx"
 
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        print(cell_obj.value)

        INF_OEE03 = cell_obj.value
        
    print(INF_OEE03.split("|"))
    
    INF_OEE03_Plant = INF_OEE03.split("|")[0]
    INF_OEE03_WC_Text = INF_OEE03.split("|")[1]
    INF_OEE03_Machine_text = INF_OEE03.split("|")[2]
    INF_OEE03_Machine_ID = INF_OEE03.split("|")[3]
    INF_OEE03_Material = INF_OEE03.split("|")[4]
    INF_OEE03_Description = INF_OEE03.split("|")[5]
    INF_OEE03_PD_Order = INF_OEE03.split("|")[6]
    INF_OEE03_Batch_No = INF_OEE03.split("|")[7]
    INF_OEE03_Tag_No = INF_OEE03.split("|")[8]
    INF_OEE03_Pallet_No = INF_OEE03.split("|")[9]
    INF_OEE03_GR_QTY_EA = INF_OEE03.split("|")[10]
    INF_OEE03_Status = INF_OEE03.split("|")[11]
    INF_OEE03_GR_Date = INF_OEE03.split("|")[12]
    INF_OEE03_GR_Time = INF_OEE03.split("|")[13]

    
    print("INF_OEE03_Plant --> " ,INF_OEE03_Plant )
    print("INF_OEE03_WC_Text --> " ,INF_OEE03_WC_Text )
    print("INF_OEE03_Machine_text --> " ,INF_OEE03_Machine_text )
    print("INF_OEE03_Machine_ID --> " ,INF_OEE03_Machine_ID )
    print("INF_OEE03_Material --> " ,INF_OEE03_Material )
    print("INF_OEE03_Description --> " ,INF_OEE03_Description )
    print("INF_OEE03_PD_Order --> " ,INF_OEE03_PD_Order )
    print("INF_OEE03_Batch_No --> " ,INF_OEE03_Batch_No )
    print("INF_OEE03_Tag_No --> " ,INF_OEE03_Tag_No )
    print("INF_OEE03_Pallet_No --> " ,INF_OEE03_Pallet_No )
    print("INF_OEE03_GR_QTY_EA --> " ,INF_OEE03_GR_QTY_EA )
    print("INF_OEE03_Status --> " ,INF_OEE03_Status )
    print("INF_OEE03_GR_Date --> " ,INF_OEE03_GR_Date )
    print("INF_OEE03_GR_Time --> " ,INF_OEE03_GR_Time )
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    INSERT_INF_OEE03 = cnxn.cursor()
    INSERT_INF_OEE03.execute('INSERT INTO OEE_DB.dbo.INF_OEE03 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,BatchNo,TagNo,PalletNo,GR_QTY_EA,Status,GRDate,GRTime) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE03_Plant  , INF_OEE03_WC_Text , INF_OEE03_Machine_text  , INF_OEE03_Machine_ID  , INF_OEE03_Material  , INF_OEE03_Description,INF_OEE03_PD_Order,INF_OEE03_Batch_No,INF_OEE03_Tag_No,INF_OEE03_Pallet_No,INF_OEE03_GR_QTY_EA,INF_OEE03_Status,INF_OEE03_GR_Date,INF_OEE03_GR_Time))
    cnxn.commit()
   
def SAP_INF_OEE04() :

    path = "D:\\Work\\foster\\oee\\web server V2.1\\web\\dataSAP\\INF_OEE04.xlsx"
 
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        print(cell_obj.value)

        INF_OEE04 = cell_obj.value
        
    print(INF_OEE04.split("|"))
    
    INF_OEE04_Plant = INF_OEE04.split("|")[0]
    INF_OEE04_WC_Text = INF_OEE04.split("|")[1]
    INF_OEE04_Machine_text = INF_OEE04.split("|")[2]
    INF_OEE04_Machine_ID = INF_OEE04.split("|")[3]
    INF_OEE04_Material = INF_OEE04.split("|")[4]
    INF_OEE04_Description = INF_OEE04.split("|")[5]
    INF_OEE04_PD_Order = INF_OEE04.split("|")[6]
    INF_OEE04_Date = INF_OEE04.split("|")[7]
    INF_OEE04_QC_Time = INF_OEE04.split("|")[8]
    INF_OEE04_QC_Qty = INF_OEE04.split("|")[9]
    INF_OEE04_GR_Dep = INF_OEE04.split("|")[10]


    
    print("INF_OEE04_Plant --> " ,INF_OEE04_Plant )
    print("INF_OEE04_WC_Text --> " ,INF_OEE04_WC_Text )
    print("INF_OEE04_Machine_text --> " ,INF_OEE04_Machine_text )
    print("INF_OEE04_Machine_ID --> " ,INF_OEE04_Machine_ID )
    print("INF_OEE04_Material --> " ,INF_OEE04_Material )
    print("INF_OEE04_Description --> " ,INF_OEE04_Description )
    print("INF_OEE04_PD_Order --> " ,INF_OEE04_PD_Order )
    print("INF_OEE04_Date --> " ,INF_OEE04_Date )
    print("INF_OEE04_QC_Time --> " ,INF_OEE04_QC_Time )
    print("INF_OEE04_QC_Qty --> " ,INF_OEE04_QC_Qty )
    print("INF_OEE04_GR_Dep --> " ,INF_OEE04_GR_Dep )
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    INSERT_INF_OEE04 = cnxn.cursor()
    INSERT_INF_OEE04.execute('INSERT INTO OEE_DB.dbo.INF_OEE04 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,Date,QCTime,QC_QTY,GR_Dep) VALUES(?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE04_Plant  , INF_OEE04_WC_Text , INF_OEE04_Machine_text  , INF_OEE04_Machine_ID  , INF_OEE04_Material  , INF_OEE04_Description,INF_OEE04_PD_Order,INF_OEE04_Date,INF_OEE04_QC_Time,INF_OEE04_QC_Qty,INF_OEE04_GR_Dep))
    cnxn.commit()


if __name__ == "__main__" :
    
    print("------------ INF_OEE01 --------------")
    SAP_INF_OEE01()
    print("------------ INF_OEE02 --------------")
    SAP_INF_OEE02()
    print("------------ INF_OEE03 --------------")
    SAP_INF_OEE03()
    print("------------ INF_OEE04 --------------")
    SAP_INF_OEE04()
    
    
    
    