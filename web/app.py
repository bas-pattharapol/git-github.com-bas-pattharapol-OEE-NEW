import email
from re import I
from flask import Flask, render_template, request, redirect, url_for, jsonify,make_response,session,send_file
from flask.sessions import NullSession
import flask_login 
from flask_login.utils import logout_user
import pandas as pd
import pyodbc
import json
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
import decimal
from openpyxl.styles import Alignment , Font , PatternFill
import openpyxl
import time 
import babel
import sys

User = "admin"
UserLevel = "1"



class Encoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, decimal.Decimal): return float(obj)
        
class create_dict(dict): 
  
    # __init__ function 
    def __init__(self): 
        self = dict() 
          
    # Function to add key:value 
    def add(self, key, value): 
        self[key] = value

server = "172.30.1.2"
port = 5432
database = "OEE_DB"
username = "sa"
password = "p@ssw0rd"

ansOEE_Plant = 'ALL'
ansOEE_Machines = 'ALL'
ansOEE_Shifts   = 'ALL'
ansOEE_StartDate = ''
ansOEE_StopDate = ''
ansOEE_UserGroup = 'ALL'

ansOEE_Plant_M = 'ALL'
ansOEE_Machines_M = 'ALL'
ansOEE_Shifts_M   = 'ALL'
ansOEE_StartDate_M = ''
ansOEE_StopDate_M = ''
ansOEE_UserGroup_M = 'ALL'
ansOEE_Month_M = 'ALL'

ansYield_Plant = 'ALL'
ansYield_Machines = 'ALL'
ansYield_Shifts   = 'ALL'
ansYield_StartDate = ''
ansYield_StopDate = ''
ansYield_UserGroup = 'ALL'


where = ''

excelOEE_Plant_M = 'ALL'
excelOEE_Machines_M = 'ALL'
excelOEE_Shifts_M   = 'ALL'
excelOEE_StartDate_M = ''
excelOEE_StopDate_M = ''
excelOEE_UserGroup_M = 'ALL'
excelOEE_Month_M = ''


excelOEE_Plant = 'ALL'
excelOEE_Machines = 'ALL'
excelOEE_Shifts   = 'ALL'
excelOEE_StartDate = ''
excelOEE_StopDate = ''
excelOEE_UserGroup = 'ALL'

ansOEE_Month_M_Report = ''

excelYield_Plant = 'ALL'
excelYield_Machines = 'ALL'
excelYield_Shifts   = 'ALL'
excelYield_StartDate = ''
excelYield_StopDate = ''
excelYield_UserGroup = 'ALL'
    

app = Flask(__name__)
app.secret_key = '0'
login_manager = flask_login.LoginManager()

login_manager.init_app(app)
users = {}
nameUser = None

count = 3

class User(flask_login.UserMixin):
    pass


def chTime(pd,ShiftCode,mode):
    
    codePlan = []
    codeUnplan = []
    count = 0
   
    
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    DownTimeCode = conn.cursor()
    DownTimeCode.execute("SELECT Code FROM OEE_DB.dbo.[DownTimeCode] WHERE DeleteFlag = 1 AND Type = 'Plan' ")
    
    for i in DownTimeCode:
        codePlan.append(i[0])
        
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    DownTimeCode = conn.cursor()
    DownTimeCode.execute("SELECT Code FROM OEE_DB.dbo.[DownTimeCode] WHERE DeleteFlag = 1 AND Type = 'Unplan' ")
    
    for i in DownTimeCode:
        codeUnplan.append(i[0])
    
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    cur = conn.cursor()
    cur.execute("SELECT StartTime , EndTime FROM OEE_DB.dbo.[ShiftCode] WHERE DeleteFlag = 1 AND ShiftCodeID = ? ",(ShiftCode,))
    
    for i in cur:
        StartTime = i[0]      
        EndTime = i[1]     
    
    if ShiftCode == '1A':
    
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT * FROM OEE_DB.dbo.[INF_OEE2_V2] WHERE DeleteFlag = 1 AND PDOrder = ? AND TypeTime = 'DonwTime' AND StartTime >= ? AND EndTime <= ? ",(pd,StartTime,EndTime))
    
    elif ShiftCode == '2A':      
          
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT * FROM OEE_DB.dbo.[INF_OEE2_V2] WHERE DeleteFlag = 1 AND PDOrder = ? AND TypeTime = 'DonwTime' AND StartTime >= ? ",(pd,StartTime))
    
    elif ShiftCode == '1B' or ShiftCode == '2B':
    
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT * FROM OEE_DB.dbo.[INF_OEE2_V2] WHERE DeleteFlag = 1 AND PDOrder = ? AND TypeTime = 'DonwTime' AND StartTime >= ? AND EndTime <= ? ",(pd,StartTime,EndTime))
    elif ShiftCode == '1OT':      
          
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT * FROM OEE_DB.dbo.[INF_OEE2_V2] WHERE DeleteFlag = 1 AND PDOrder = ? AND TypeTime = 'DonwTime' AND StartTime >= ? AND EndTime <= ? ",(pd,StartTime,EndTime))
    
    elif ShiftCode == '2OT':      
          
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT Min , DownTimeCode FROM OEE_DB.dbo.[INF_OEE2_V2] WHERE DeleteFlag = 1 AND PDOrder = ? AND TypeTime = 'DonwTime' AND (StartTime >= ? OR StartTime <= ? ) AND (EndTime <= ? OR EndTime >= ? ) ",(pd,StartTime,EndTime,EndTime,StartTime))
    
    if mode == 'Plan' :
        for k in cur:
            if k[1] in codePlan:
                count += int(k[0])//60
                
    elif mode == 'Unplan':
        for k in cur:
            if k[1] in codeUnplan:
                count += int(k[0])//60
          
    return count
            
@login_manager.user_loader
def user_loader(email):
    if email not in users:
        return

    user = User()
    user.id = email
    return user

@login_manager.request_loader
def request_loader(request):
    email = request.form.get('Username')
    if email not in users:
        return

    user = User()
    user.id = email
    return user
    
@app.errorhandler(401)
def custom_401(error):
    return redirect(url_for('login_user'))

@app.before_request
def before_request():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(minutes=30)


@app.template_filter()
def format_datetime(value, format='medium'):
    if format == 'full':
        format="EEEE, d. MMMM y 'at' HH:mm"
    elif format == 'medium':
        format="EE dd.MM.y HH:mm"
    return babel.dates.format_datetime(value, format)

@app.route("/API_INF_OEE01", methods=['POST'])
def API_INF_OEE01():
    print(request.get_json())
    data = request.get_json()

    for i in range(0,len(data['Result'])):
        print('Plant --> ' ,data['Result'][i]['Plant'] )
        print('PDOrder --> ' ,data['Result'][i]['PDOrder'] )
        print('MachineID --> ' ,data['Result'][i]['MachineID'] )
        print('Material --> ' ,data['Result'][i]['Material'] )
        print('Description --> ' ,data['Result'][i]['Description'] )
        print('PlanQuantity --> ' ,data['Result'][i]['PlanQuantity'] )
        print('Bacth --> ' ,data['Result'][i]['Bacth'] )
        print('Bulk - Code --> ' ,data['Result'][i]['Code'] )
        print('Bulk - PD_order1 --> ' ,data['Result'][i]['PD_order1'] )
        print('Bulk - PD_order2 --> ' ,data['Result'][i]['PD_order2'] )
        print('Bulk - PD_order3 --> ' ,data['Result'][i]['PD_order3'] )
        print('Bulk - PD_order4 --> ' ,data['Result'][i]['PD_order4'] )
        print('Bulk - PD_order5 --> ' ,data['Result'][i]['PD_order5'] )
        print('Bulk - PD_order6 --> ' ,data['Result'][i]['PD_order6'] )
        print('Bulk - PD_order7 --> ' ,data['Result'][i]['PD_order7'] )
        print('Bulk - PD_order8 --> ' ,data['Result'][i]['PD_order8'] )
        print('Bulk - PD_order9 --> ' ,data['Result'][i]['PD_order9'] )
        print('Bulk - PD_order10 --> ' ,data['Result'][i]['PD_order10'] )
        
        print("------------------------------------")
        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.INF_OEE1_V2 (Plant, PDOrder, MachineID, Material, Description, PlanQuantity, Bacth, Code, PD_order1, PD_order2, PD_order3, PD_order4, PD_order5, PD_order6, PD_order7, PD_order8) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,(data['Result'][i]['Plant'],data['Result'][i]['PDOrder'],data['Result'][i]['MachineID'],data['Result'][i]['Material'],data['Result'][i]['Description'],data['Result'][i]['PlanQuantity'],data['Result'][i]['Bacth'],data['Result'][i]['Code'],data['Result'][i]['PD_order1'],data['Result'][i]['PD_order2'],data['Result'][i]['PD_order3'],data['Result'][i]['PD_order4'],data['Result'][i]['PD_order5'],data['Result'][i]['PD_order6'],data['Result'][i]['PD_order7'],data['Result'][i]['PD_order8']))
        cnxn.commit()
        
    return json.dumps({'success':True}), 200, {'ContentType':'application/json'} 

@app.route("/API_INF_OEE03", methods=['POST'])
def API_INF_OEE03():
    print(request.get_json())
    data = request.get_json()
    for p in range(0,len(data['Result'])):
        for i in range(0,len(data['Result'][p]['Machine'])):
        
            for j in range(0,len(data['Result'][p]['Machine'][i]['GR_QTY'])):
                print('PDOrder --> ' ,data['Result'][p]['PDOrder'] )
                print('Machine - ID --> ' ,data['Result'][p]['Machine'][i]['ID'] )
                print('GR_QTY - ID --> ' ,data['Result'][p]['Machine'][i]['GR_QTY'][j]['QTY'] )
                print('GR_QTY - Date --> ' ,data['Result'][p]['Machine'][i]['GR_QTY'][j]['Date'] )
                print('GR_QTY - Time --> ' ,data['Result'][p]['Machine'][i]['GR_QTY'][j]['Time'] )
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                update = cnxn.cursor()
                update.execute('INSERT INTO OEE_DB.dbo.INF_OEE3_V2 (PDOrder,MachineID, QTY, [Date], [Time]) VALUES(?,?,?,?,?)' ,(data['Result'][p]['PDOrder'],data['Result'][p]['Machine'][i]['ID'],data['Result'][p]['Machine'][i]['GR_QTY'][j]['QTY'],data['Result'][p]['Machine'][i]['GR_QTY'][j]['Date'],data['Result'][p]['Machine'][i]['GR_QTY'][j]['Time']))
                cnxn.commit()
            print("------------------------------------")
        
    
    return json.dumps({'success':True}), 200, {'ContentType':'application/json'} 

@app.route("/API_INF_OEE04", methods=['POST'])
def API_INF_OEE04(): 
    print(request.get_json())
    data = request.get_json()
    for p in range(0,len(data['Result'])):
        for i in range(0,len(data['Result'][p]['Machine'])):
            print('PDOrder --> ' ,data['Result'][p]['PDOrder'] )
            print('Machine - ID --> ' ,data['Result'][p]['Machine'][i]['ID'] )
            for j in range(0,len(data['Result'][p]['Machine'][i]['GI'])):
                print('GI - ID --> ' ,data['Result'][p]['Machine'][i]['GI'][j]['QTY'] )
                print('GI - Dep --> ' ,data['Result'][p]['Machine'][i]['GI'][j]['Dep'] )
                print('GI - Date --> ' ,data['Result'][p]['Machine'][i]['GI'][j]['Date'] )
                print('GI - Time --> ' ,data['Result'][p]['Machine'][i]['GI'][j]['Time'] )
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                update = cnxn.cursor()
                update.execute('INSERT INTO OEE_DB.dbo.INF_OEE4_V2 (PDOrder,MachineID, QTY,Dep ,[Date], [Time]) VALUES(?,?,?,?,?,?)' ,(data['Result'][p]['PDOrder'],data['Result'][p]['Machine'][i]['ID'],data['Result'][p]['Machine'][i]['QC'][j]['QTY'],data['Result'][p]['Machine'][i]['QC'][j]['Dep'],data['Result'][p]['Machine'][i]['QC'][j]['Date'],data['Result'][p]['Machine'][i]['QC'][j]['Time']))
                cnxn.commit()
            print("------------------------------------")
        
    return json.dumps({'success':True}), 200, {'ContentType':'application/json'} 

@app.route("/API_INF_OEE05", methods=['POST'])
def API_INF_OEE05():
    print(request.get_json())
    data = request.get_json()
    for p in range(0,len(data['Result'])):
        for i in range(0,len(data['Result'][p]['Machine'])):
            print('PDOrder --> ' ,data['Result'][p]['PDOrder'] )
            print('Machine - ID --> ' ,data['Result'][p]['Machine'][i]['ID'] )
            for j in range(0,len(data['Result'][p]['Machine'][i]['Return'])):
                print('Return - ID --> ' ,data['Result'][p]['Machine'][i]['Return'][j]['QTY'] )
                print('Return - Text --> ' ,data['Result'][p]['Machine'][i]['Return'][j]['Text'] )
                print('Return - Date --> ' ,data['Result'][p]['Machine'][i]['Return'][j]['Date'] )
                print('Return - Time --> ' ,data['Result'][p]['Machine'][i]['Return'][j]['Time'] )
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                update = cnxn.cursor()
                update.execute('INSERT INTO OEE_DB.dbo.INF_OEE5_V2 (PDOrder,MachineID, QTY,Text ,[Date], [Time]) VALUES(?,?,?,?,?,?)' ,(data['Result'][p]['PDOrder'],data['Result'][p]['Machine'][i]['ID'],data['Result'][p]['Machine'][i]['Return'][j]['QTY'],data['Result'][p]['Machine'][i]['Return'][j]['Text'],data['Result'][p]['Machine'][i]['Return'][j]['Date'],data['Result'][p]['Machine'][i]['Return'][j]['Time']))
                cnxn.commit()
            print("------------------------------------")
        
    return json.dumps({'success':True}), 200, {'ContentType':'application/json'} 

@app.route("/API_RunTime_DownTime", methods=['POST'])
def API_RunTime_DownTime():
    print(request.get_json())
    data = request.get_json()
    

    for i in range(0,len(data['RunTime'])):
        print("-->> RunTime [",i,"]")
        print('PDOrder --> ' ,data['PDOrder'] )
        print('RunTime - BatchNo --> ' ,data['RunTime'][i]['BatchNo'] )
        print('RunTime - PostDate --> ' ,data['RunTime'][i]['PostDate'] )
        print('RunTime - Shift --> ' ,data['RunTime'][i]['Shift'] )
        print('RunTime - StartTime --> ' ,data['RunTime'][i]['StartTime'] )
        print('RunTime - EndTime --> ' ,data['RunTime'][i]['EndTime'] )
        print('RunTime - Time --> ' ,data['RunTime'][i]['Time'] )
        print("------------------------------------")
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.INF_OEE2_V2 (PDOrder, TypeTime, BatchNo, PostDate, Shift, StartTime, EndTime, [Min]) VALUES(?,?,?,?,?,?,?,?)' ,(data['PDOrder'],"RunTime",data['RunTime'][i]['BatchNo'],data['RunTime'][i]['PostDate'],data['RunTime'][i]['Shift'],data['RunTime'][i]['StartTime'],data['RunTime'][i]['EndTime'],data['RunTime'][i]['Time']))
        cnxn.commit()
        
    for i in range(0,len(data['DonwTime'])):
        print("-->> DonwTime [",i,"]")
        print('PDOrder --> ' ,data['PDOrder'] )
        print('DonwTime - BatchNo --> ' ,data['DonwTime'][i]['BatchNo'] )
        print('DonwTime - PostDate --> ' ,data['DonwTime'][i]['PostDate'] )
        print('DonwTime - Shift --> ' ,data['DonwTime'][i]['Shift'] )
        print('DonwTime - StartTime --> ' ,data['DonwTime'][i]['StartTime'] )
        print('DonwTime - EndTime --> ' ,data['DonwTime'][i]['EndTime'] )
        print('DonwTime - DownTimeCode --> ' ,data['DonwTime'][i]['DownTimeCode'] )
        print('DonwTime - EndTime --> ' ,data['DonwTime'][i]['EndTime'] )
        print('DonwTime - Time --> ' ,data['DonwTime'][i]['Time'] )
                    
        print("------------------------------------")
        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.INF_OEE2_V2 (PDOrder, TypeTime, BatchNo, PostDate, Shift, StartTime, EndTime, [Min],DownTimeCode) VALUES(?,?,?,?,?,?,?,?,?)' ,(data['PDOrder'],"DonwTime",data['DonwTime'][i]['BatchNo'],data['DonwTime'][i]['PostDate'],data['DonwTime'][i]['Shift'],data['DonwTime'][i]['StartTime'],data['DonwTime'][i]['EndTime'],data['DonwTime'][i]['Time'],data['DonwTime'][i]['DownTimeCode']))
        cnxn.commit()
    
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    cur1 = conn.cursor()
    cur1.execute(""" SELECT TOP(1) iov.MachineID , ppt.PlannedCode from OEE_DB.dbo.INF_OEE1_V2 iov
                    INNER JOIN OEE_DB.dbo.PlannedProductionTime ppt
                    ON iov.MachineID = ppt.MachineID AND iov.PDOrder = ? AND ppt.[Date] = ? 
                    order by ppt.[DateTime] DESC 
                """,(data['PDOrder'] ,data['RunTime'][0]['PostDate']))
    
    startTime = []
    endTime = []   
    
    for l in cur1 :
        print(l)
        
        if l[1] == 'AA':
            print('1A Plan', chTime(data['PDOrder'],'1A','Plan'))
            print('1A Unplan', chTime(data['PDOrder'],'1A','Unplan'))
            print('2A Plan', chTime(data['PDOrder'],'2A','Plan'))
            print('2A Unplan', chTime(data['PDOrder'],'2A','Unplan'))
        elif l[1] == 'BB':
            print('1B Plan', chTime(data['PDOrder'],'1B','Plan'))
            print('1B Unplan', chTime(data['PDOrder'],'1B','Unplan'))
            print('2B Plan', chTime(data['PDOrder'],'2B','Plan'))
            print('2B Unplan', chTime(data['PDOrder'],'2B','Unplan'))
        elif l[1] == 'TOT':  
            print('1OT Plan', chTime(data['PDOrder'],'1OT','Plan'))
            print('1OT Unplan', chTime(data['PDOrder'],'1OT','Unplan'))
            print('2OT Plan', chTime(data['PDOrder'],'2OT','Plan'))
            print('2OT Unplan', chTime(data['PDOrder'],'2OT','Unplan'))
        elif l[1] == 'HOT': 
            chTime(data['PDOrder'],'1A','Plan')
        else:
            print(l[1] ,"Plan" , chTime(data['PDOrder'],l[1],'Plan'))
            print(l[1] ,"Unplan", chTime(data['PDOrder'],l[1],'Unplan'))
                #chTime('1A','2A',data['DonwTime'][i]['StartTime'],data['DonwTime'][i]['EndTime'] ,data['DonwTime'][i]['DownTimeCode'] )
            
                
            print(startTime)
            print(endTime)
            
        
        
        
    return json.dumps({'success':True}), 200, {'ContentType':'application/json'} 

@app.route("/", methods=['GET', 'POST'])
@app.route("/login_user", methods=['GET', 'POST'])
def login_user():        
    
    return render_template('login.html')

@app.route("/sidebar", methods=['GET', 'POST'])
def sidebar():        
    return render_template('sidebar.html')

@app.route("/login_CkPass", methods=['GET', 'POST'])
def login_CkPass():
    if request.method == 'POST':
        
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT * FROM OEE_DB.dbo.[User] WHERE DeleteFlag = 1")

        dataUser = request.form['username']
        print(users)
        for data in cur: 
            print(data)
            if dataUser == data[3]:
                if data[4] == None or  data[4] == '':
                    return render_template('login_NoPass.html',username = data[3])        
                else:
                    return render_template('login_Password.html',username = data[3])   
           
            
        return redirect(url_for('login_user'))

@app.route("/login_addPass/<string:username1>", methods=['GET', 'POST'])
def login_addPass(username1):
    if request.method == 'POST':
        dataPassword = request.form['New_Password']
        NewPassword = generate_password_hash(dataPassword,"sha256")
        print(NewPassword)
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID=sa ;PWD=p@ssw0rd')
        cur = conn.cursor()
        cur.execute("UPDATE OEE_DB.dbo.[User] SET Pass = ? WHERE UserName = ?", (NewPassword,username1))
        conn.commit() 
       

    return redirect(url_for('login_user'))

@app.route("/login_Password/<string:username1>", methods=['GET', 'POST'])
def login_Password(username1):
    if request.method == 'POST':
        dataPassword = request.form['Password']

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute("SELECT * FROM OEE_DB.dbo.[User] WHERE UserName = ?  AND DeleteFlag = 1", (username1,))

        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur1 = conn.cursor()
        cur1.execute("SELECT * FROM OEE_DB.dbo.[User]  WHERE DeleteFlag = 1")

        email = username1

        for i in cur1:
            users[i[3]] = {'Password' : i[4]}

        print(users)
        
        
        for data in cur: 
            print(data[3])
            Level = data[11]
            Fname_LName = data[5] + ' ' + data[6]
            print(check_password_hash(data[4], dataPassword))
           
            if check_password_hash(data[4], dataPassword):
                user = User()
                user.id = email
                flask_login.login_user(user)
                    
                global nameUser 
                nameUser = email
                return render_template('login_succeed.html',username = username1,Level=Level,Fname_LName = Fname_LName)   
            else :
                return redirect(url_for('login_user'))

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login_user'))

@app.errorhandler(401)
def custom_401(error):
    return redirect(url_for('login_user'))

@app.errorhandler(404)
def custom_404(error):
    return redirect(url_for('login_user'))


@app.errorhandler(500)
def custom_500(error):
    return redirect(url_for('login_user'))


@app.route('/userManagement/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def userManagement(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
           

            UserName = request.form['UserName']
            Fname = request.form['Fname']
            Lname = request.form['Lname']
            UserGroup = request.form['UserGroup']
            UserLevel = request.form['UserLevel']

            Department = request.form['Department']
            
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            userManagement = cnxn.cursor()
            userManagement.execute('UPDATE OEE_DB.dbo.[User] SET UserName = ? ,Fname = ? , Lname = ?, UserGroup = ? , UserLevel = ? , UserLevelID = ?, Department = ? ,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ? ',(UserName,Fname,Lname,UserGroup,UserLevel,UserLevel,Department,'1',id ))
            cnxn.commit()
            
        elif mode == "add":

            UserName = request.form['UserName']
            Fname = request.form['Fname']
            Lname = request.form['Lname']
            UserGroup = request.form['UserGroup']
            UserLevel = request.form['UserLevel']

            Department = request.form['Department']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            userManagement = cnxn.cursor()
            userManagement.execute('INSERT INTO OEE_DB.dbo.[User] ( UserName ,Fname , Lname , UserGroup  , UserLevel , UserLevelID, Department,DeleteFlag  ) VALUES(?,?,?,?,?,?,?,?)' ,(UserName,Fname,Lname,UserGroup,UserLevel,UserLevel,Department,'1') )
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(8,mode,id,Fname_Lname,Level))
        cnxn.commit()
            
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        userManagement = cnxn.cursor()
        userManagement.execute('UPDATE OEE_DB.dbo.[User] SET DeleteFlag = ?  , DateTime = GETDATE() WHERE RecordID = ? ',(-1,id) )
        cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(8,mode,id,Fname_Lname,Level))
        cnxn.commit()
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    userManagement = cnxn.cursor()
    userManagement.execute('SELECT * FROM OEE_DB.dbo.[User] WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Department = cnxn.cursor()
    Department.execute('SELECT DepartmentID FROM OEE_DB.dbo.Department WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    lenDepartment = 0
    Department_s = []
    for i in Department:
        lenDepartment+=1
        Department_s.append(i)

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    UserLevel = cnxn.cursor()
    UserLevel.execute('SELECT * FROM OEE_DB.dbo.UserLevel ')
    lenUserLevel = 0
    UserLevel_s = []
    for i in UserLevel:
        lenUserLevel+=1
        UserLevel_s.append(i)
    
        
    return render_template('userManagement.html',userManagement = userManagement,Department=Department_s,lenDepartment=lenDepartment,UserLevel=UserLevel_s,lenUserLevel=lenUserLevel,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/index/<string:username1>/<string:Level>/<string:Fname_Lname>')
@flask_login.login_required
def index(username1,Level,Fname_Lname):
    global count
    count+=1
    print(count)
    return render_template('index.html', username =username1, Level = Level,Fname_Lname = Fname_Lname)

@app.route("/uploadFile/<string:Level>/<string:Fname_Lname>", methods=['GET', 'POST'])
@flask_login.login_required
def uploadFile(Level,Fname_Lname):
    if request.method == 'POST':
        print(request.files['file'])
        #print(request.files['file2'])
        
        f = request.files['file']
        col = ['Plant','MachineID','Machine','Loading','Sat','Sun','Mon','Tue','Wed','The','Fri','Sat1','Sun1']
        data_xls = pd.read_excel(f,usecols ='A:M',skiprows=2,header=None,names=col)
        data_xls1 = pd.read_excel(f,usecols ='A:M',skiprows=3,header=None,names=col)
        dimensions = data_xls.shape
        print(dimensions[0])
        dataPlant = []
        dataMachineID = []
        dataMachine =[]
        dataLoading=[]
        dataPlannedCode=[]
        dataDate=[]
        dataStart=[]
        dataEnd=[]
            
        dataGet = ['Sat','Sun','Mon','Tue','Wed','The','Fri','Sat1','Sun1']
        now = datetime.now()
        #now = datetime(2021, 5, 17)
      
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute(' SELECT ShiftCodeID ,StartTime ,EndTime FROM OEE_DB.dbo.ShiftCode ')
        Plant_s = []
        PlannedCode_Chk = []
        StartTime = []
        EndTime = []
        PlantID_Chk = []
        PlantName_Chk = []
        dataPlantName= []
        len_Plant_s = 0
        for i in cur:
            len_Plant_s+=1
            Plant_s.append(i)
                    
        for i in range(0,len_Plant_s):
           PlannedCode_Chk.append(Plant_s[i][0])
               
        for i in range(0,len_Plant_s):
           StartTime.append(Plant_s[i][1].strftime("%H:%M:%S"))
               
        for i in range(0,len_Plant_s):
           EndTime.append(Plant_s[i][2].strftime("%H:%M:%S"))
               
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        cur.execute('SELECT PlantID ,PlantName FROM OEE_DB.dbo.Plant WHERE DeleteFlag = 1')
            
        len_Plant_data = 0
        Plant_data = []
        for i in cur:
            len_Plant_data+=1
            Plant_data.append(i)
            
        for i in range(0,len_Plant_data):
           PlantID_Chk.append(Plant_data[i][0])
               
        for i in range(0,len_Plant_data):
           PlantName_Chk.append(Plant_data[i][1])
        
            
        print(PlannedCode_Chk)
        print(PlantName_Chk)
        print(StartTime)
        print(EndTime)
        for i in range(0,int(dimensions[0])-1):
            for j in range(0,9):
                
                if data_xls[dataGet[j]][0] >= now :
                
                    dataPlant.append(data_xls1['Plant'][i])
                    dataMachineID.append(data_xls1['MachineID'][i])
                    dataMachine.append(data_xls1['Machine'][i])
                    dataLoading.append(data_xls1['Loading'][i])
                    dataPlannedCode.append(data_xls1[dataGet[j]][i])
                        
                    for o in range(0,len_Plant_s):
                        if data_xls1[dataGet[j]][i] == PlannedCode_Chk[o] :
                            dataStart.append(StartTime[o])
                            dataEnd.append(EndTime[o])
                                
                    for o in range(0,len_Plant_data):
                        if PlantName_Chk[o] == data_xls1['Plant'][i] :
                            dataPlantName.append(PlantID_Chk[o])
                           
                        
                    dataDate.append(data_xls[dataGet[j]][0].date())
       
        ok_data  = pd.DataFrame({'PlantName':dataPlant,
                                 'PlantID':dataPlantName,
                                    'MachineID':dataMachineID,
                                    'Machine':dataMachine,
                                    'Loading':dataLoading,
                                    'Date':dataDate,
                                    'PlannedCode':dataPlannedCode,
                                    'StartTime':dataStart,
                                    'EndTime':dataEnd})
        print(ok_data)
        conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        cur = conn.cursor()
        for row in ok_data.itertuples():
            postgres_insert_query = ' INSERT INTO OEE_DB.dbo.PlannedProductionTime (PlantID,PlantName,MachineID,MachineName,LoadingDate,Date,PlannedCode,StartTime,EndTime,DeleteFlag) VALUES (?,?,?,?,?,?,?,?,?,?)'
            record_to_insert = (row.PlantID,row.PlantName, row.MachineID, row.Machine,row.Loading,row.Date,row.PlannedCode,row.StartTime,row.EndTime,'1')
            cur.execute(postgres_insert_query, record_to_insert)

        conn.commit()
            
        return redirect(url_for('uploadFile',Level=Level,Fname_Lname=Fname_Lname))
      
    return render_template('uploadFile.html',Level=Level,Fname_Lname=Fname_Lname)

@app.route("/testDataExcel_INF_OEE/<string:io>", methods=['GET', 'POST'])
def testDataExcel_INF_OEE(io):
    if request.method == 'POST':
        if io == '01':
            print(request.files['file'])
            #print(request.files['file2'])
            
            f = request.files['file']
            col = ['Plant','WC_Text','Machine_Text','Machine_ID','Material','Description','PD_Order','UOM_Qty','Plan_quantity','Batch','Bulk_Code','Bulk_PD_order1','Bulk_PD_order2','Bulk_PD_order3','Bulk_PD_order4','Bulk_PD_order5','Bulk_PD_order6','Bulk_PD_order7','Bulk_PD_order8','Bulk_PD_order9','Bulk_PD_order10']
            data_xls = pd.read_excel(f,names=col)
            dimensions = data_xls.shape
            print(data_xls)
            
            
            for i in range(0,dimensions[0]):
                #Plant.append(data_xls['Plant'][i])
                INF_OEE01_Plant = str(data_xls['Plant'][i])
                INF_OEE01_WC_Text = str(data_xls['WC_Text'][i])
                INF_OEE01_Machine_text = str(data_xls['Machine_Text'][i])
                INF_OEE01_Machine_ID = str(data_xls['Machine_ID'][i])
                INF_OEE01_Material = str(data_xls['Material'][i])
                INF_OEE01_Description = str(data_xls['Description'][i])
                INF_OEE01_PD_Order = str(data_xls['PD_Order'][i])
                INF_OEE01_UOM_Qty = str(data_xls['UOM_Qty'][i])
                INF_OEE01_Plan_quantity = str(data_xls['Plan_quantity'][i])
                INF_OEE01_Batch = str(data_xls['Batch'][i])
                INF_OEE01_Bulk_Code = str(data_xls['Bulk_Code'][i])
                INF_OEE01_Bulk_PD_order1 = str(data_xls['Bulk_PD_order1'][i])
                INF_OEE01_Bulk_PD_order2 = str(data_xls['Bulk_PD_order2'][i])
                INF_OEE01_Bulk_PD_order3 = str(data_xls['Bulk_PD_order3'][i])
                INF_OEE01_Bulk_PD_order4 = str(data_xls['Bulk_PD_order4'][i])
                INF_OEE01_Bulk_PD_order5 = str(data_xls['Bulk_PD_order5'][i])
                INF_OEE01_Bulk_PD_order6 = str(data_xls['Bulk_PD_order6'][i])
                INF_OEE01_Bulk_PD_order7 = str(data_xls['Bulk_PD_order7'][i])
                INF_OEE01_Bulk_PD_order8 = str(data_xls['Bulk_PD_order8'][i])
                INF_OEE01_Bulk_PD_order9 = str(data_xls['Bulk_PD_order9'][i])
                INF_OEE01_Bulk_PD_order10 = str(data_xls['Bulk_PD_order10'][i])
                print("INF_OEE01_Plant --> " ,INF_OEE01_Plant )
                print("INF_OEE01_WC_Text --> " ,INF_OEE01_WC_Text )
                print("INF_OEE01_Machine_text --> " ,INF_OEE01_Machine_text )
                print("INF_OEE01_Machine_ID --> " ,INF_OEE01_Machine_ID )
                print("INF_OEE01_Material --> " ,INF_OEE01_Material )
                print("INF_OEE01_Description --> " ,INF_OEE01_Description )
                print("INF_OEE01_PD_Order --> " ,INF_OEE01_PD_Order )
                print("INF_OEE01_UOM_Qty --> " ,INF_OEE01_UOM_Qty )
                print("INF_OEE01_Plan_quantity --> " ,INF_OEE01_Plan_quantity )
                print("INF_OEE01_Batch --> " ,INF_OEE01_Batch )
                print("INF_OEE01_Bulk_Code --> " ,INF_OEE01_Bulk_Code )
                print("INF_OEE01_Bulk_PD_order1 --> " ,INF_OEE01_Bulk_PD_order1 )
                print("INF_OEE01_Bulk_PD_order2 --> " ,INF_OEE01_Bulk_PD_order2 )
                print("INF_OEE01_Bulk_PD_order3 --> " ,INF_OEE01_Bulk_PD_order3 )
                print("INF_OEE01_Bulk_PD_order4 --> " ,INF_OEE01_Bulk_PD_order4 )
                print("INF_OEE01_Bulk_PD_order5 --> " ,INF_OEE01_Bulk_PD_order5 )
                print("INF_OEE01_Bulk_PD_order6 --> " ,INF_OEE01_Bulk_PD_order6 )
                print("INF_OEE01_Bulk_PD_order7 --> " ,INF_OEE01_Bulk_PD_order7 )
                print("INF_OEE01_Bulk_PD_order8 --> " ,INF_OEE01_Bulk_PD_order8 )
                print("INF_OEE01_Bulk_PD_order9 --> " ,INF_OEE01_Bulk_PD_order9 )
                print("INF_OEE01_Bulk_PD_order10 --> " ,INF_OEE01_Bulk_PD_order10 )
            
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                INSERT_INF_OEE01 = cnxn.cursor()
                INSERT_INF_OEE01.execute('INSERT INTO OEE_DB.dbo.INF_OEE01 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,UOMQty,PlanQuantity,Batch,BulkCode,BulkPDOrder1,BulkPDOrder2,BulkPDOrder3,BulkPDOrder4,BulkPDOrder5,BulkPDOrder6,BulkPDOrder7,BulkPDOrder8,BulkPDOrder9,BulkPDOrder10) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE01_Plant  , INF_OEE01_WC_Text , INF_OEE01_Machine_text  , INF_OEE01_Machine_ID  , INF_OEE01_Material  , INF_OEE01_Description,INF_OEE01_PD_Order,INF_OEE01_UOM_Qty,INF_OEE01_Plan_quantity,INF_OEE01_Batch,INF_OEE01_Bulk_Code,INF_OEE01_Bulk_PD_order1,INF_OEE01_Bulk_PD_order2,INF_OEE01_Bulk_PD_order3,INF_OEE01_Bulk_PD_order4,INF_OEE01_Bulk_PD_order5,INF_OEE01_Bulk_PD_order6,INF_OEE01_Bulk_PD_order7,INF_OEE01_Bulk_PD_order8,INF_OEE01_Bulk_PD_order9,INF_OEE01_Bulk_PD_order10 ))
                cnxn.commit()
            

        elif io == '02':
            print(request.files['file'])
            #print(request.files['file2'])
            
            f = request.files['file']
            col = ['Plant','WC_Text','Machine_Text','Machine_ID','Material','Description','PD_Order','UOM_Qty','Plan_quantity','Batch_Estimate','Down_time_code_1','Down_time_description_1','Hrs_1','Down_time_code_2','Down_time_description_2','Hrs_2','Down_time_code_3','Down_time_description_3','Hrs_3','Down_time_code_4','Down_time_description_4','Hrs_4','Down_time_code_5','Down_time_description_5','Hrs_5']
            data_xls = pd.read_excel(f,names=col)
            dimensions = data_xls.shape
            print(data_xls)
            
            
            for i in range(0,dimensions[0]):
                #Plant.append(data_xls['Plant'][i])
                INF_OEE02_Plant = str(data_xls['Plant'][i])
                INF_OEE02_WC_Text =str(data_xls['WC_Text'][i])
                INF_OEE02_Machine_text = str(data_xls['Machine_Text'][i])
                INF_OEE02_Machine_ID = str(data_xls['Machine_ID'][i])
                INF_OEE02_Material = str(data_xls['Material'][i])
                INF_OEE02_Description = str(data_xls['Description'][i])
                INF_OEE02_PD_Order = str(data_xls['PD_Order'][i])
                INF_OEE02_UOM_Qty = str(data_xls['UOM_Qty'][i])
                INF_OEE02_Plan_quantity = str(data_xls['Plan_quantity'][i])
                INF_OEE02_Batch_Estimate = str(data_xls['Batch_Estimate'][i])
                INF_OEE02_Down_time_code1  = str(data_xls['Down_time_code_1'][i])
                INF_OEE02_Down_time_description1 = str(data_xls['Down_time_description_1'][i])
                INF_OEE02_Hrs1 = str(data_xls['Hrs_1'][i])
                INF_OEE02_Down_time_code2  = str(data_xls['Down_time_code_2'][i])
                INF_OEE02_Down_time_description2= str(data_xls['Down_time_description_2'][i])
                INF_OEE02_Hrs2 = str(data_xls['Hrs_2'][i])
                INF_OEE02_Down_time_code3  = str(data_xls['Down_time_code_3'][i])
                INF_OEE02_Down_time_description3 = str(data_xls['Down_time_description_3'][i])
                INF_OEE02_Hrs3 = str(data_xls['Hrs_3'][i])
                INF_OEE02_Down_time_code4  = str(data_xls['Down_time_code_4'][i])
                INF_OEE02_Down_time_description4 = str(data_xls['Down_time_description_4'][i])
                INF_OEE02_Hrs4 = str(data_xls['Hrs_4'][i])
                INF_OEE02_Down_time_code5  = str(data_xls['Down_time_code_5'][i])
                INF_OEE02_Down_time_description5 = str(data_xls['Down_time_description_5'][i])
                INF_OEE02_Hrs5 = str(data_xls['Hrs_5'][i])
                

                
                print("INF_OEE02_Plant --> " ,INF_OEE02_Plant )
                print("INF_OEE02_WC_Text --> " ,INF_OEE02_WC_Text )
                print("INF_OEE02_Machine_text --> " ,INF_OEE02_Machine_text )
                print("INF_OEE02_Machine_ID --> " ,INF_OEE02_Machine_ID )
                print("INF_OEE02_Material --> " ,INF_OEE02_Material )
                print("INF_OEE02_Description --> " ,INF_OEE02_Description )
                print("INF_OEE02_PD_Order --> " ,INF_OEE02_PD_Order )
                print("INF_OEE02_UOM_Qty --> " ,INF_OEE02_UOM_Qty )
                print("INF_OEE02_Plan_quantity --> " ,INF_OEE02_Plan_quantity )
                print("INF_OEE02_Batch_Estimate --> " ,INF_OEE02_Batch_Estimate )
                print("INF_OEE02_Down_time_code1 --> " ,INF_OEE02_Down_time_code1 )
                print("INF_OEE02_Down_time_description1 --> " ,INF_OEE02_Down_time_description1 )
                print("INF_OEE02_Hrs1 --> " ,INF_OEE02_Hrs1 )
                print("INF_OEE02_Down_time_code2 --> " ,INF_OEE02_Down_time_code2 )
                print("INF_OEE02_Down_time_description2 --> " ,INF_OEE02_Down_time_description2 )
                print("INF_OEE02_Hrs2 --> " ,INF_OEE02_Hrs2 )
                print("INF_OEE02_Down_time_code3 --> " ,INF_OEE02_Down_time_code3 )
                print("INF_OEE02_Down_time_description3 --> " ,INF_OEE02_Down_time_description3 )
                print("INF_OEE02_Hrs3 --> " ,INF_OEE02_Hrs3 )
                print("INF_OEE02_Down_time_code4 --> " ,INF_OEE02_Down_time_code4 )
                print("INF_OEE02_Down_time_description4 --> " ,INF_OEE02_Down_time_description4 )
                print("INF_OEE02_Hrs4 --> " ,INF_OEE02_Hrs4 )
                print("INF_OEE02_Down_time_code5 --> " ,INF_OEE02_Down_time_code5 )
                print("INF_OEE02_Down_time_description5 --> " ,INF_OEE02_Down_time_description5 )
                print("INF_OEE02_Hrs5 --> " ,INF_OEE02_Hrs5 )
                
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                INSERT_INF_OEE02 = cnxn.cursor()
                INSERT_INF_OEE02.execute('INSERT INTO OEE_DB.dbo.INF_OEE02 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,UOMQty,PlanQuantity,BatchEstimate,DownTimeCode1,DownTimeDescription1,Hrs1,DownTimeCode2,DownTimeDescription2,Hrs2,DownTimeCode3,DownTimeDescription3,Hrs3,DownTimeCode4,DownTimeDescription4,Hrs4,DownTimeCode5,DownTimeDescription5,Hrs5) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE02_Plant  , INF_OEE02_WC_Text , INF_OEE02_Machine_text  , INF_OEE02_Machine_ID  , INF_OEE02_Material  , INF_OEE02_Description,INF_OEE02_PD_Order,INF_OEE02_UOM_Qty,INF_OEE02_Plan_quantity,INF_OEE02_Batch_Estimate,INF_OEE02_Down_time_code1,INF_OEE02_Down_time_description1,INF_OEE02_Hrs1,INF_OEE02_Down_time_code2,INF_OEE02_Down_time_description2,INF_OEE02_Hrs2,INF_OEE02_Down_time_code3,INF_OEE02_Down_time_description3,INF_OEE02_Hrs3,INF_OEE02_Down_time_code4,INF_OEE02_Down_time_description4,INF_OEE02_Hrs4,INF_OEE02_Down_time_code5,INF_OEE02_Down_time_description5,INF_OEE02_Hrs5))
                cnxn.commit()
            
        elif io == '03':
            print(request.files['file'])
            #print(request.files['file2'])
            
            f = request.files['file']
            col = ['Plant','WC_Text','Machine_Text','Machine_ID','Material','Description','PD_Order','Batch_No','Tag_No','Pallet_No','GR_QTY','Status','GR_Date','GR_Time']
            data_xls = pd.read_excel(f,names=col)
            dimensions = data_xls.shape
            print(data_xls)
            
            for i in range(0,dimensions[0]):
                #Plant.append(data_xls['Plant'][i])
                INF_OEE03_Plant = str(data_xls['Plant'][i])
                INF_OEE03_WC_Text =str(data_xls['WC_Text'][i])
                INF_OEE03_Machine_text = str(data_xls['Machine_Text'][i])
                INF_OEE03_Machine_ID = str(data_xls['Machine_ID'][i])
                INF_OEE03_Material = str(data_xls['Material'][i])
                INF_OEE03_Description = str(data_xls['Description'][i])
                INF_OEE03_PD_Order = str(data_xls['PD_Order'][i])
                INF_OEE03_Batch_No = str(data_xls['Batch_No'][i])
                INF_OEE03_Tag_No = str(data_xls['Tag_No'][i])
                INF_OEE03_Pallet_No = str(data_xls['Pallet_No'][i])
                INF_OEE03_GR_QTY = str(data_xls['GR_QTY'][i])
                INF_OEE03_Status = str(data_xls['Status'][i])
                INF_OEE03_GR_Date = str(data_xls['GR_Date'][i])
                INF_OEE03_GR_Time = str(data_xls['GR_Time'][i])
                
                print("INF_OEE03_Plant --> " ,INF_OEE03_Plant )
                print("INF_OEE03_WC_Text --> " ,INF_OEE03_WC_Text )
                print("INF_OEE03_Machine_text --> " ,INF_OEE03_Machine_text )
                print("INF_OEE03_Machine_ID --> " ,INF_OEE03_Machine_ID )
                print("INF_OEE03_Material --> " ,INF_OEE03_Material )
                print("INF_OEE03_Description --> " ,INF_OEE03_Description )
                print("INF_OEE03_PD_Order --> " ,INF_OEE03_PD_Order )
                print("INF_OEE03_Batch_No --> " ,INF_OEE03_Batch_No )
                print("INF_OEE03_Tag_No --> " ,INF_OEE03_Tag_No )
                print("INF_OEE03_Pallet_No --> " ,INF_OEE03_Pallet_No )
                print("INF_OEE03_GR_QTY --> " ,INF_OEE03_GR_QTY )
                print("INF_OEE03_Status --> " ,INF_OEE03_Status )
                print("INF_OEE03_GR_Date --> " ,INF_OEE03_GR_Date )
                print("INF_OEE03_GR_Time --> " ,INF_OEE03_GR_Time )
                
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                INSERT_INF_OEE03 = cnxn.cursor()
                INSERT_INF_OEE03.execute('INSERT INTO OEE_DB.dbo.INF_OEE03 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,BatchNo,TagNo,PalletNo,GR_QTY_EA,Status,GRDate,GRTime) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE03_Plant  , INF_OEE03_WC_Text , INF_OEE03_Machine_text  , INF_OEE03_Machine_ID  , INF_OEE03_Material  , INF_OEE03_Description,INF_OEE03_PD_Order,INF_OEE03_Batch_No,INF_OEE03_Tag_No,INF_OEE03_Pallet_No,INF_OEE03_GR_QTY,INF_OEE03_Status,INF_OEE03_GR_Date,INF_OEE03_GR_Time))
                cnxn.commit()
                
        elif io == '04':
            print(request.files['file'])
            #print(request.files['file2'])
            
            f = request.files['file']
            col = ['Plant','WC_Text','Machine_Text','Machine_ID','Material','Description','PD_Order','Date','QC_time','QC_Qty','GR_Dep']
            data_xls = pd.read_excel(f,names=col)
            dimensions = data_xls.shape
            print(data_xls)
            
            for i in range(0,dimensions[0]):
                #Plant.append(data_xls['Plant'][i])
                INF_OEE04_Plant = str(data_xls['Plant'][i])
                INF_OEE04_WC_Text =str(data_xls['WC_Text'][i])
                INF_OEE04_Machine_text = str(data_xls['Machine_Text'][i])
                INF_OEE04_Machine_ID = str(data_xls['Machine_ID'][i])
                INF_OEE04_Material = str(data_xls['Material'][i])
                INF_OEE04_Description = str(data_xls['Description'][i])
                INF_OEE04_PD_Order = str(data_xls['PD_Order'][i])
                INF_OEE04_Date = str(data_xls['Date'][i])
                INF_OEE04_QC_time = str(data_xls['QC_time'][i])
                INF_OEE04_QC_Qty = str(data_xls['QC_Qty'][i])
                INF_OEE04_GR_Dep = str(data_xls['GR_Dep'][i])
               
                
                print("INF_OEE03_Plant --> " ,INF_OEE04_Plant )
                print("INF_OEE03_WC_Text --> " ,INF_OEE04_WC_Text )
                print("INF_OEE03_Machine_text --> " ,INF_OEE04_Machine_text )
                print("INF_OEE03_Machine_ID --> " ,INF_OEE04_Machine_ID )
                print("INF_OEE03_Material --> " ,INF_OEE04_Material )
                print("INF_OEE03_Description --> " ,INF_OEE04_Description )
                print("INF_OEE03_PD_Order --> " ,INF_OEE04_PD_Order )
                print("INF_OEE04_Date --> " ,INF_OEE04_Date )
                print("INF_OEE04_QC_time --> " ,INF_OEE04_QC_time )
                print("INF_OEE04_QC_Qty --> " ,INF_OEE04_QC_Qty )
                print("INF_OEE04_GR_Dep --> " ,INF_OEE04_GR_Dep )
                
                
                cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
                INSERT_INF_OEE04 = cnxn.cursor()
                INSERT_INF_OEE04.execute('INSERT INTO OEE_DB.dbo.INF_OEE04 (Plant,W_CText,Machine_Text,MachineID,Material,Description,PDOrder,Date,QCTime,QC_QTY,GR_Dep) VALUES(?,?,?,?,?,?,?,?,?,?,?)' ,( INF_OEE04_Plant  , INF_OEE04_WC_Text , INF_OEE04_Machine_text  , INF_OEE04_Machine_ID  , INF_OEE04_Material  , INF_OEE04_Description,INF_OEE04_PD_Order,INF_OEE04_Date,INF_OEE04_QC_time,INF_OEE04_QC_Qty,INF_OEE04_GR_Dep))
                cnxn.commit()
        else:
            io = "None - Error"
      
    return render_template('testDataExcel.html',io=io)

@app.route('/Data_API') 
def Data_API():
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    workMachine = cnxn.cursor()
    workMachine.execute("""SELECT RecordID ,DateTime , PlantName,MachineID,MachineName,PlannedCode,Date,StartTime,EndTime FROM OEE_DB.dbo.PlannedProductionTime  AS Q
                            WHERE DateTime >= (
                                SELECT MAX(DateTime)
                                FROM OEE_DB.dbo.PlannedProductionTime
                                WHERE MachineID = Q.MachineID AND Date = Q.Date
                            )

                            order by Date asc """)
    payload = []
    content = {}
    for result in workMachine:
        content = {'PlantName': result[2], 'MachineID': str(result[3]), 'MachineName': result[4],'PlannedCode': result[5],'Date': str(result[6]),'StartTime': str(result[7]),'EndTime': str(result[8])}
        payload.append(content)
        content = {}
    #print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201


@app.route('/DataTableShift')
@flask_login.login_required
def DataTableShift():
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    MachineID = cnxn.cursor()
    MachineID.execute('SELECT MachineID FROM OEE_DB.dbo.Machines WHERE DeleteFlag = 1')
    MachineID_s = []
    len_Machine = 0
    for i in MachineID:
        len_Machine+=1
        MachineID_s.append(i)
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    MachineName = cnxn.cursor()
    MachineName.execute('SELECT MachineName FROM OEE_DB.dbo.Machines WHERE DeleteFlag = 1')
    
    MachineName_s = []
    for i in MachineName:
        MachineName_s.append(i)
       
    print(MachineID_s)         
    print(MachineName_s)
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    workMachine = cnxn.cursor()
    workMachine.execute("""SELECT RecordID ,DateTime , PlantName,MachineID,MachineName,PlannedCode,Date,StartTime,EndTime FROM OEE_DB.dbo.PlannedProductionTime  AS Q
                            WHERE DateTime >= (
                                SELECT MAX(DateTime)
                                FROM OEE_DB.dbo.PlannedProductionTime
                                WHERE MachineID = Q.MachineID AND Date = Q.Date
                            )
                            order by Date asc """)

    
    
    return render_template('DataTableShift.html',Machines = workMachine,MachineID_s= MachineID_s,MachineName_s=MachineName_s,len_Machine=len_Machine)

@app.route('/dashboard')
@flask_login.login_required
def dashboard():
    return render_template('Metrics.html')

@app.route('/oee_Total')
@flask_login.login_required
def oee_Total():
    return render_template('oee_Total.html')

@app.route('/oee_MachineV2')
@flask_login.login_required
def oee_MachineV2():
    return render_template('oee_MachineV2.html')

@app.route('/oee_miru')
@flask_login.login_required
def oee_miru():
    return render_template('oee_miru.html')

@app.route('/yield_Machine')
@flask_login.login_required
def yield_Machine():
    return render_template('yield_Machine.html')

@app.route('/yield_total')
@flask_login.login_required
def yield_total():
    return render_template('yield_total.html')

@app.route('/Edit_StorageTanks/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_StorageTanks(mode,id,Level,Fname_Lname):
    global User,UserLevel

    if request.method == 'POST':
        if mode == 'update':
            StorageTanksID = request.form['StorageTanksID']
            StorageTanksName = request.form['StorageTanksName']
            StorageTanksDesc = request.form['StorageTanksDesc']
            PlantName = request.form['PlantName']
            Material = request.form['Material']
            MainProduct = request.form['MainProduct']
            SubProduct = "-"
            SetTime = request.form['SetTime']
            ValidatedSpeed = request.form['ValidatedSpeed']
            MaxSpeed = request.form['MaxSpeed'] 
             
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            StorageTanks = cnxn.cursor()
            StorageTanks.execute(' UPDATE OEE_DB.dbo.StorageTanks  SET MachineID = ? , MachineName = ?,MachineDesc = ?,PlantName = ?,Material = ?,MainProduct = ?,SubProduct = ?,ValidatedSpeed = ?,SetTime = ?,MaxSpeed = ?,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ? ', ( StorageTanksID  , StorageTanksName , StorageTanksDesc , PlantName ,Material ,MainProduct ,SubProduct,ValidatedSpeed,SetTime,MaxSpeed ,'1',id ))
            cnxn.commit()
        
       

        elif mode == "add":
            StorageTanksID = request.form['StorageTanksID']
            StorageTanksName = request.form['StorageTanksName']
            StorageTanksDesc = request.form['StorageTanksDesc']
            PlantName = request.form['PlantName']
            Material = request.form['Material']
            MainProduct = request.form['MainProduct']
            SubProduct = "-"
            SetTime = request.form['SetTime']
            ValidatedSpeed = request.form['ValidatedSpeed']
            MaxSpeed = request.form['MaxSpeed']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            StorageTanks = cnxn.cursor()
            StorageTanks.execute(' INSERT INTO OEE_DB.dbo.StorageTanks  (MachineID,MachineName,MachineDesc,PlantName ,Material, MainProduct,SubProduct,ValidatedSpeed,SetTime,MaxSpeed,DeleteFlag) VALUES(?,?,?,?,?,?,?,?,?,?,?)' , ( StorageTanksID  , StorageTanksName , StorageTanksDesc , PlantName ,Material, MainProduct ,SubProduct,ValidatedSpeed,SetTime,MaxSpeed ,'1' ))
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(2,mode,StorageTanksName,Fname_Lname,Level))
        cnxn.commit()  
            
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        Machines = cnxn.cursor()
        Machines.execute('UPDATE OEE_DB.dbo.StorageTanks  SET DeleteFlag = -1 , DateTime = GETDATE() WHERE RecordID = ? ',id )
        cnxn.commit()   

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(2,mode,id,Fname_Lname,Level))
        cnxn.commit() 
    
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    StorageTanks = cnxn.cursor()
    StorageTanks.execute('SELECT * FROM OEE_DB.dbo.StorageTanks  WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT PlantName FROM OEE_DB.dbo.Plant WHERE DeleteFlag = 1 ')
    lenPlant = 0
    Plant_s = []
    for i in Plant:
        lenPlant+=1
        Plant_s.append(i)
        
    print(Plant_s)
        
    return render_template('Edit_StorageTanks.html',StorageTanks = StorageTanks,Plant=Plant_s,len=lenPlant,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_Machines/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_Machines(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            MachineID = request.form['MachineID']
            MachineName = request.form['MachineName']
            MachineDesc = request.form['MachineDesc']
            PlantName = request.form['PlantName']
            Material = request.form['Material']
            MainProduct = request.form['MainProduct']
            SubProduct = "-"
            SetTime = request.form['SetTime']
            ValidatedSpeed = request.form['ValidatedSpeed']
            MaxSpeed = request.form['MaxSpeed'] 
             
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            StorageTanks = cnxn.cursor()
            StorageTanks.execute(' UPDATE OEE_DB.dbo.Machines SET MachineID = ? , MachineName = ?,MachineDesc = ?,PlantName = ?,Material = ?,MainProduct = ?,SubProduct = ?,ValidatedSpeed = ?,SetTime = ?,MaxSpeed = ?,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ? ', ( MachineID  , MachineName , MachineDesc , PlantName ,Material ,MainProduct ,SubProduct,ValidatedSpeed,SetTime,MaxSpeed ,'1',id ))
            cnxn.commit()
        
       

        elif mode == "add":
            MachineID = request.form['MachineID']
            MachineName = request.form['MachineName']
            MachineDesc = request.form['MachineDesc']
            PlantName = request.form['PlantName']
            Material = request.form['Material']
            MainProduct = request.form['MainProduct']
            SubProduct = "-"
            SetTime = request.form['SetTime']
            ValidatedSpeed = request.form['ValidatedSpeed']
            MaxSpeed = request.form['MaxSpeed']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            StorageTanks = cnxn.cursor()
            StorageTanks.execute(' INSERT INTO OEE_DB.dbo.Machines (MachineID,MachineName,MachineDesc,PlantName ,Material, MainProduct,SubProduct,ValidatedSpeed,SetTime,MaxSpeed,DeleteFlag) VALUES(?,?,?,?,?,?,?,?,?,?,?)' , ( MachineID  , MachineName , MachineDesc , PlantName ,Material, MainProduct ,SubProduct,ValidatedSpeed,SetTime,MaxSpeed ,'1' ))
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(2,mode,MachineName,Fname_Lname,Level))
        cnxn.commit()  
            
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        Machines = cnxn.cursor()
        Machines.execute('UPDATE OEE_DB.dbo.Machines SET DeleteFlag = -1 , DateTime = GETDATE() WHERE RecordID = ? ',id )
        cnxn.commit()   

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(2,mode,id,Fname_Lname,Level))
        cnxn.commit() 
    
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Machines = cnxn.cursor()
    Machines.execute('SELECT * FROM OEE_DB.dbo.Machines WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT PlantName FROM OEE_DB.dbo.Plant WHERE DeleteFlag = 1 ')
    lenPlant = 0
    Plant_s = []
    for i in Plant:
        lenPlant+=1
        Plant_s.append(i)
        
    print(Plant_s)
        
    return render_template('Edit_Machines.html',Machines = Machines,Plant=Plant_s,len = lenPlant,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_Plant/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_Plant(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            PlantID = request.form['PlantID']
            PlantName = request.form['PlantName']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Plant = cnxn.cursor()
            Plant.execute(' UPDATE OEE_DB.dbo.Plant SET PlantID = ? , PlantName = ?,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ?', (  PlantID  , PlantName , '1',id ))
            cnxn.commit()
            
        elif mode == "add":
            PlantID = request.form['PlantID']
            PlantName = request.form['PlantName']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Plant = cnxn.cursor()
            Plant.execute(' INSERT INTO OEE_DB.dbo.Plant (PlantID,PlantName,DeleteFlag) VALUES(?,?,?)' ,( PlantID  , PlantName , '1' ))
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(3,mode,PlantName,Fname_Lname,Level))
        cnxn.commit() 
            
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        Plant = cnxn.cursor()
        Plant.execute('UPDATE OEE_DB.dbo.Plant SET DeleteFlag = -1 , DateTime = GETDATE() WHERE RecordID =? ',id )
        cnxn.commit()
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(3,mode,id,Fname_Lname,Level))
        cnxn.commit() 
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT * FROM OEE_DB.dbo.Plant WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    return render_template('Edit_Plant.html',Plant = Plant,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_DownTimeCode/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_DownTimeCode(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            Code = request.form['Code']
            Description = request.form['Description']
            Type = request.form['Type']
            Response = request.form['Response']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            DownTimeCode = cnxn.cursor()
            DownTimeCode.execute('UPDATE OEE_DB.dbo.DownTimeCode SET Code = ?,Description = ?,Type = ?,Response = ?,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID =? ', ( Code,Description ,Type,Response, '1',id ))
            cnxn.commit()
            
        elif mode == "add":
            
            Code = request.form['Code']
            Description = request.form['Description']
            Type = request.form['Type']
            Response = request.form['Response']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            DownTimeCode = cnxn.cursor()
            DownTimeCode.execute(' INSERT INTO OEE_DB.dbo.DownTimeCode (Code,Description ,Type,Response,DeleteFlag) VALUES(?,?,?,?,?)' , ( Code,Description ,Type,Response, '1' ))
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(4,mode,Description,Fname_Lname,Level))
        cnxn.commit()      
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        DownTimeCode = cnxn.cursor()
        DownTimeCode.execute(' UPDATE OEE_DB.dbo.DownTimeCode SET DeleteFlag = -1 , DateTime = GETDATE() WHERE  RecordID = ? ',id )
        cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(4,mode,id,Fname_Lname,Level))
        cnxn.commit() 
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    DownTimeCode = cnxn.cursor()
    DownTimeCode.execute('SELECT * FROM OEE_DB.dbo.DownTimeCode WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Department = cnxn.cursor()
    Department.execute('SELECT DepartmentName FROM OEE_DB.dbo.Department WHERE DeleteFlag = 1 ')
    lenDepartment = 0
    Department_s = []
    for i in Department:
        lenDepartment+=1
        Department_s.append(i)
        
    #print(Department_s)
        
    return render_template('Edit_DownTimeCode.html',DownTimeCode = DownTimeCode,Department=Department_s,len = lenDepartment,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_Shift/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_Shift(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            ShiftCodeID = request.form['ShiftCodeID']
            ShiftCodeName = request.form['ShiftCodeName']
            StartTime = request.form['StartTime']
            EndTime = request.form['EndTime']
            Break1 = request.form['Break1']
            Break2 = request.form['Break2']
            Break3 = request.form['Break3']
            PlanProductionTime = request.form['PlanProductionTime']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Edit_Shift = cnxn.cursor() 
            Edit_Shift.execute('UPDATE OEE_DB.dbo.ShiftCode SET ShiftCodeID = ? , ShiftCodeName = ?,StartTime = ?,EndTime = ?,Break1 = ?,Break2 = ?,Break3 = ?,PlanProductionTime = ?,DeleteFlag = ?  , DateTime = GETDATE() WHERE RecordID = ? ',( ShiftCodeID  , ShiftCodeName,StartTime ,EndTime,Break1,Break2,Break3,PlanProductionTime ,'1',id ))
            cnxn.commit()
            
        elif mode == "add":
            ShiftCodeID = request.form['ShiftCodeID']
            ShiftCodeName = request.form['ShiftCodeName']
            StartTime = request.form['StartTime']
            EndTime = request.form['EndTime']
            Break1 = request.form['Break1']
            Break2 = request.form['Break2']
            Break3 = request.form['Break3']
            PlanProductionTime = request.form['PlanProductionTime']
            
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Edit_Shift = cnxn.cursor()
            Edit_Shift.execute('INSERT INTO OEE_DB.dbo.ShiftCode (ShiftCodeID , ShiftCodeName,StartTime,EndTime,Break1,Break2,Break3,PlanProductionTime,DeleteFlag) VALUES(?,?,?,?,?,?,?,?,?)' ,(  ShiftCodeID  , ShiftCodeName,StartTime ,EndTime,Break1,Break2,Break3,PlanProductionTime, '1') )
            cnxn.commit()
        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(5,mode,ShiftCodeName,Fname_Lname,Level))
        cnxn.commit() 
            
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        Edit_Shift = cnxn.cursor()
        Edit_Shift.execute('UPDATE OEE_DB.dbo.ShiftCode SET DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ? ',(-1,id) )
        cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(5,mode,id,Fname_Lname,Level))
        cnxn.commit() 
        
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Edit_Shift = cnxn.cursor()
    Edit_Shift.execute('SELECT * FROM OEE_DB.dbo.ShiftCode WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    
        
    return render_template('Edit_Shift.html',Edit_Shift = Edit_Shift,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_Standardcolor/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_Standardcolor(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            Red = request.form['Red']
            Yellow = request.form['Yellow']
            Green = request.form['Green']
            Revision = request.form['Revision']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            DownTimeCode = cnxn.cursor()
            DownTimeCode.execute('UPDATE OEE_DB.dbo.StandardColor SET Red = ? , Yellow = ?,Green = ?,Revision = ?,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ? ', (Red  , Yellow ,Green ,Revision, '1',id ))
            cnxn.commit()
            
        elif mode == "add":
            Red = request.form['Red']
            Yellow = request.form['Yellow']
            Green = request.form['Green']
            Revision = request.form['Revision']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            DownTimeCode = cnxn.cursor()
            DownTimeCode.execute('INSERT INTO StandardColor (Red ,Yellow ,Green ,Revision,DeleteFlag) VALUES(?,?,?,?,?)',( Red  , Yellow ,Green ,Revision, '1' ))
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(6,mode,Revision,Fname_Lname,Level))
        cnxn.commit()     
            
   
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Edit_Standardcolor = cnxn.cursor()
    Edit_Standardcolor.execute('SELECT * FROM OEE_DB.dbo.StandardColor WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    Standardcolor = []
    
    for i in Edit_Standardcolor:
    
        #len_color+=1
        Standardcolor.append(i)
    
        
    return render_template('Edit_Standardcolor.html',Edit_Standardcolor = Standardcolor,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_UserGroup/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_UserGroup(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            UserGroupID = request.form['UserGroupID']
            UserGroupName = request.form['UserGroupName']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Edit_UserGroup = cnxn.cursor()
            Edit_UserGroup.execute('UPDATE OEE_DB.dbo.UserGroup SET UserGroupID = ? , UserGroupName = ? ,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID = ? ',(UserGroupID,UserGroupName  ,'1',id ))
            cnxn.commit()
            
        elif mode == "add":
            UserGroupID = request.form['UserGroupID']
            UserGroupName = request.form['UserGroupName']
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Edit_UserGroup = cnxn.cursor()
            Edit_UserGroup.execute('INSERT INTO OEE_DB.dbo.UserGroup (UserGroupID , UserGroupName,DeleteFlag) VALUES(?,?,?)' ,( UserGroupID,UserGroupName  ,'1') )
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(7,mode,UserGroupName,Fname_Lname,Level))
        cnxn.commit()
            
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        Edit_UserGroup = cnxn.cursor()
        Edit_UserGroup.execute('UPDATE OEE_DB.dbo.UserGroup SET DeleteFlag = ?  , DateTime = GETDATE() WHERE RecordID = ? ',(-1,id) )
        cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(7,mode,id,Fname_Lname,Level))
        cnxn.commit()
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Edit_UserGroup = cnxn.cursor()
    Edit_UserGroup.execute('SELECT * FROM OEE_DB.dbo.UserGroup WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    
        
    return render_template('Edit_UserGroup.html',Edit_UserGroup = Edit_UserGroup,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/Edit_Department/<string:mode>/<string:id>/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def Edit_Department(mode,id,Level,Fname_Lname):
    if request.method == 'POST':
        if mode == 'update':
            Department_ID = request.form['Department_ID']
            Department_Name = request.form['Department_Name']
           
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Department = cnxn.cursor()
            Department.execute('UPDATE OEE_DB.dbo.Department SET DepartmentID = ?,DepartmentName = ?,DeleteFlag = ? , DateTime = GETDATE() WHERE RecordID =? ', ( Department_ID,Department_Name ,'1',id ))
            cnxn.commit()
            
        elif mode == "add":
            
            Department_ID = request.form['Department_ID']
            Department_Name = request.form['Department_Name']
           
            
            cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
            Department = cnxn.cursor()
            Department.execute(' INSERT INTO OEE_DB.dbo.Department (DepartmentID,DepartmentName ,DeleteFlag) VALUES(?,?,?)' , (Department_ID,Department_Name , '1' ))
            cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(9,mode,Department_Name,Fname_Lname,Level))
        cnxn.commit()      
    if mode == "del":
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        Department = cnxn.cursor()
        Department.execute(' UPDATE OEE_DB.dbo.Department SET DeleteFlag = -1 , DateTime = GETDATE() WHERE  RecordID = ? ',id )
        cnxn.commit()

        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        update = cnxn.cursor()
        update.execute('INSERT INTO OEE_DB.dbo.EditRecord (EditID,ChangeTopic,NameValue,[User],UserLevel) VALUES(?,?,?,?,?)' ,(9,mode,id,Fname_Lname,Level))
        cnxn.commit() 
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Department = cnxn.cursor()
    Department.execute('SELECT * FROM OEE_DB.dbo.Department WHERE DeleteFlag = 1 ORDER BY DateTime DESC')
    
    
        
    return render_template('Edit_Department.html',Department = Department,Level=Level,Fname_Lname=Fname_Lname)


@app.route('/hello')
@flask_login.login_required
def hello():
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    workMachine = cnxn.cursor()
    workMachine.execute('SELECT RecordID, PlantID , PlantName,MachineID,MachineName,Date,PlannedCode,StartTime,EndTime FROM OEE_DB.dbo.PlannedProductionTime  ORDER BY DateTime DESC ')

    record = workMachine.fetchone()
    
    return record

@app.route('/Record/<string:Level>/<string:Fname_Lname>')
@flask_login.login_required
def Record(Level,Fname_Lname):
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Record = cnxn.cursor()
    Record.execute('SELECT * FROM OEE_DB.dbo.EditRecord INNER JOIN OEE_DB.dbo.EditCode ON EditRecord.EditID=EditCode.ID ORDER BY DateTime DESC')

    
    
    return render_template('Record.html',Record = Record,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/ReportOEE/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def ReportOEE(Level,Fname_Lname):
    global ansOEE_Plant
    global ansOEE_Machines
    global ansOEE_Shifts
    global ansOEE_StartDate
    global ansOEE_StopDate
    global where
    global ansOEE_UserGroup
    
    global excelOEE_Plant
    global excelOEE_Machines 
    global excelOEE_Shifts   
    global excelOEE_StartDate 
    global excelOEE_StopDate 
    global excelOEE_UserGroup 
    
    ansOEE_Plant = 'ALL'
    ansOEE_Machines = 'ALL'
    ansOEE_Shifts   = 'ALL'
    ansOEE_UserGroup = 'ALL'
    ansOEE_StartDate = ''
    ansOEE_StopDate = ''
    
    
    if request.method == 'POST':
        ansOEE_Plant = request.form['Plant']
        ansOEE_Machines = request.form['Machines']
        ansOEE_Shifts = request.form['Shifts']
        ansOEE_StartDate = request.form['StartDate']
        ansOEE_StopDate = request.form['StopDate']
        ansOEE_UserGroup = request.form['UserGroup']
        print("---------------")
        print(ansOEE_Plant)
        print(ansOEE_Machines)
        print(ansOEE_Shifts)
        print(ansOEE_StartDate)
        print(ansOEE_StopDate)
        print(ansOEE_UserGroup)
        
        
        excelOEE_Plant = ansOEE_Plant
        excelOEE_Machines = ansOEE_Machines
        excelOEE_Shifts   = ansOEE_Shifts
        excelOEE_StartDate = ansOEE_StartDate
        excelOEE_StopDate = ansOEE_StopDate
        excelOEE_UserGroup = ansOEE_UserGroup
    
   
    if ansOEE_Plant == 'ALL':
        ansOEE_Plant = ''
    else : 
        ansOEE_Plant = " AND PlantName = '"+ ansOEE_Plant +"'"
        
    if ansOEE_Machines == 'ALL':
        ansOEE_Machines = ''
    else :  
        ansOEE_Machines = " AND MachineID = '"+ ansOEE_Machines +"'"
       
    if ansOEE_Shifts == 'ALL':
        ansOEE_Shifts = ''
    else : 
        ansOEE_Shifts = " AND ShiftCode = '"+ ansOEE_Shifts +"'"
        
    if ansOEE_UserGroup == 'ALL':
        ansOEE_UserGroup = ''
    else : 
        ansOEE_UserGroup = " AND UserGroupID = '"+ ansOEE_UserGroup +"'"
        
    if ansOEE_StartDate == '':
        ansOEE_StartDate = ''
    else : 
        ansOEE_StartDate = " AND CONVERT(DATE, DateTime) >= '"+ ansOEE_StartDate +"'"
         
    if ansOEE_StopDate == '':
        ansOEE_StopDate = ''
    else : 
        ansOEE_StopDate = " AND CONVERT(DATE, DateTime)  <= '"+ ansOEE_StopDate +"'"
      
  
        
    print("---------------")
    print(ansOEE_Plant)
    print(ansOEE_Machines)
    print(ansOEE_Shifts)
    print(ansOEE_StartDate)
    print(ansOEE_StopDate)
     
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT PlantName FROM OEE_DB.dbo.Plant  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Machines = cnxn.cursor()
    Machines.execute('SELECT MachineName ,MachineID FROM OEE_DB.dbo.Machines  WHERE DeleteFlag = 1')

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Shifts = cnxn.cursor()
    Shifts.execute('SELECT ShiftCodeName , ShiftCodeID FROM OEE_DB.dbo.ShiftCode  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    UserGroup = cnxn.cursor()
    UserGroup.execute('SELECT UserGroupName , UserGroupID FROM OEE_DB.dbo.UserGroup  WHERE DeleteFlag = 1')

    return render_template('Report_OEE.html',Plant = Plant , Machines = Machines , Shifts = Shifts,UserGroup=UserGroup,Level=Level,Fname_Lname=Fname_Lname)

@app.route('/ReportYield/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def ReportYield(Level,Fname_Lname):
    global ansYield_Plant
    global ansYield_Machines
    global ansYield_Shifts
    global ansYield_StartDate
    global ansYield_StopDate
    global where
    global ansYield_UserGroup
    
    global excelYield_Plant
    global excelYield_Machines 
    global excelYield_Shifts   
    global excelYield_StartDate 
    global excelYield_StopDate 
    global excelYield_UserGroup 
    
    ansYield_Plant = 'ALL'
    ansYield_Machines = 'ALL'
    ansYield_Shifts   = 'ALL'
    ansYield_UserGroup = 'ALL'
    ansYield_StartDate = ''
    ansYield_StopDate = ''
    
    
    if request.method == 'POST':
        ansYield_Plant = request.form['Plant']
        ansYield_Machines = request.form['Machines']
        ansYield_Shifts = request.form['Shifts']
        ansYield_StartDate = request.form['StartDate']
        ansYield_StopDate = request.form['StopDate']
        ansYield_UserGroup = request.form['UserGroup']
        print("---------------")
        print(ansYield_Plant)
        print(ansYield_Machines)
        print(ansYield_Shifts)
        print(ansYield_StartDate)
        print(ansYield_StopDate)
        print(ansYield_UserGroup)
        
        
        excelYield_Plant = ansYield_Plant
        excelYield_Machines = ansYield_Machines
        excelYield_Shifts   = ansYield_Shifts
        excelYield_StartDate = ansYield_StartDate
        excelYield_StopDate = ansYield_StopDate
        excelYield_UserGroup = ansYield_UserGroup
    
   
    if ansYield_Plant == 'ALL':
        ansYield_Plant = ''
    else : 
        ansYield_Plant = " AND PlantName = '"+ ansYield_Plant +"'"
        
    if ansYield_Machines == 'ALL':
        ansYield_Machines = ''
    else :  
        ansYield_Machines = " AND MachineID = '"+ ansYield_Machines +"'"
       
    if ansYield_Shifts == 'ALL':
        ansYield_Shifts = ''
    else : 
        ansYield_Shifts = " AND ShiftCode = '"+ ansYield_Shifts +"'"
        
    if ansYield_UserGroup == 'ALL':
        ansYield_UserGroup = ''
    else : 
        ansYield_UserGroup = " AND UserGroupID = '"+ ansYield_UserGroup +"'"
        
    if ansYield_StartDate == '':
        ansYield_StartDate = ''
    else : 
        ansYield_StartDate = " AND CONVERT(DATE, DateTime) >= '"+ ansYield_StartDate +"'"
         
    if ansYield_StopDate == '':
        ansYield_StopDate = ''
    else : 
        ansYield_StopDate = " AND CONVERT(DATE, DateTime)  <= '"+ ansYield_StopDate +"'"
      
  
        
    print("---------------")
    print(ansYield_Plant)
    print(ansYield_Machines)
    print(ansYield_Shifts)
    print(ansYield_StartDate)
    print(ansYield_StopDate)
     
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT PlantName FROM OEE_DB.dbo.Plant  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Machines = cnxn.cursor()
    Machines.execute('SELECT MachineName ,MachineID FROM OEE_DB.dbo.Machines  WHERE DeleteFlag = 1')

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Shifts = cnxn.cursor()
    Shifts.execute('SELECT ShiftCodeName , ShiftCodeID FROM OEE_DB.dbo.ShiftCode  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    UserGroup = cnxn.cursor()
    UserGroup.execute('SELECT UserGroupName , UserGroupID FROM OEE_DB.dbo.UserGroup  WHERE DeleteFlag = 1')

    return render_template('Report_Yield.html',Plant = Plant , Machines = Machines , Shifts = Shifts,UserGroup=UserGroup,Level=Level,Fname_Lname=Fname_Lname)

    
@app.route('/ReportOverallYield/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def ReportOverallYield(Level,Fname_Lname):
    return render_template('Report_Overall_Yield.html',Level=Level,Fname_Lname=Fname_Lname)

@app.route('/ReportOEEMontly/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def ReportOEEMontly(Level,Fname_Lname):
    global ansOEE_Plant_M
    global ansOEE_Machines_M
    global ansOEE_Shifts_M
    global ansOEE_StartDate_M
    global ansOEE_StopDate_M
    global where
    global ansOEE_UserGroup_M
    global ansOEE_Month_M
    global ansOEE_Month_M_Report
    
    global excelOEE_Plant_M
    global excelOEE_Machines_M 
    global excelOEE_Shifts_M   
    global excelOEE_StartDate_M 
    global excelOEE_StopDate_M 
    global excelOEE_UserGroup_M
    global excelOEE_Month_M
    ansOEE_Plant_M = 'ALL'
    ansOEE_Machines_M = 'ALL'
    ansOEE_Shifts_M   = 'ALL'
    ansOEE_UserGroup_M = 'ALL'
    ansOEE_StartDate_M = ''
    ansOEE_StopDate_M = ''
    ansOEE_Month_M = ''
    
    
    if request.method == 'POST':
        ansOEE_Plant_M = request.form['Plant']
        ansOEE_Machines_M = request.form['Machines']
        ansOEE_Shifts_M = request.form['Shifts']
        ansOEE_Month_M = request.form['Month12']
        ansOEE_UserGroup_M = request.form['UserGroup']
        ansOEE_Month_M_Report = ansOEE_Month_M
        ansOEE_Month_M = ansOEE_Month_M.replace("-", "")
     
        print("---------------")
        print(ansOEE_Plant_M)
        print(ansOEE_Machines_M)
        print(ansOEE_Shifts_M)
        print(ansOEE_StartDate_M)
        print(ansOEE_StopDate_M)
        print(ansOEE_UserGroup_M)
        print(ansOEE_Month_M)
        
        
        excelOEE_Plant_M = ansOEE_Plant_M
        excelOEE_Machines_M = ansOEE_Machines_M
        excelOEE_Shifts_M   = ansOEE_Shifts_M
        excelOEE_Month_M = ansOEE_Month_M_Report
       
        excelOEE_UserGroup_M = ansOEE_UserGroup_M
    
   
    if ansOEE_Plant_M == 'ALL':
        ansOEE_Plant_M = ''
    else : 
        ansOEE_Plant_M = " AND PlantName = '"+ ansOEE_Plant_M +"'"
        
    if ansOEE_Machines_M == 'ALL':
        ansOEE_Machines_M = ''
    else :  
        ansOEE_Machines_M = " AND MachineID = '"+ ansOEE_Machines_M +"'"
       
    if ansOEE_Shifts_M == 'ALL':
        ansOEE_Shifts_M = ''
    else : 
        ansOEE_Shifts_M = " AND ShiftCode = '"+ ansOEE_Shifts_M +"'"
        
    if ansOEE_UserGroup_M == 'ALL':
        ansOEE_UserGroup_M = ''
    else : 
        ansOEE_UserGroup_M = " AND UserGroupID = '"+ ansOEE_UserGroup_M +"'"
        
    if ansOEE_Month_M == '':
        ansOEE_Month_M = ''
    else : 
        ansOEE_Month_M = " AND CAST(FORMAT(PostingDate, 'yyyyMM') AS int) = " + ansOEE_Month_M 
         
        
    print("---------------")
    print(ansOEE_Plant_M)
    print(ansOEE_Machines_M)
    print(ansOEE_Shifts_M)
    print(ansOEE_StartDate_M)
    print(ansOEE_StopDate_M)
    print(ansOEE_Month_M)
     
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT PlantName FROM OEE_DB.dbo.Plant  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Machines = cnxn.cursor()
    Machines.execute('SELECT MachineName ,MachineID FROM OEE_DB.dbo.Machines  WHERE DeleteFlag = 1')

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Shifts = cnxn.cursor()
    Shifts.execute('SELECT ShiftCodeName , ShiftCodeID FROM OEE_DB.dbo.ShiftCode  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    UserGroup = cnxn.cursor()
    UserGroup.execute('SELECT UserGroupName , UserGroupID FROM OEE_DB.dbo.UserGroup  WHERE DeleteFlag = 1')

    return render_template('Report_OEE_Montly.html',Plant = Plant , Machines = Machines , Shifts = Shifts,UserGroup=UserGroup,Level=Level,Fname_Lname=Fname_Lname)


@app.route('/ReportYieldMontly/<string:Level>/<string:Fname_Lname>',methods=['GET', 'POST'])
@flask_login.login_required
def ReportYieldMontly(Level,Fname_Lname):
    global ansYield_Plant_M
    global ansYield_Machines_M
    global ansYield_Shifts_M
    global ansYield_StartDate_M
    global ansYield_StopDate_M
    global where
    global ansYield_UserGroup_M
    global ansYield_Month_M
    global ansYield_Month_M_Report
    
    global excelYield_Plant_M
    global excelYield_Machines_M 
    global excelYield_Shifts_M   
    global excelYield_StartDate_M 
    global excelYield_StopDate_M 
    global excelYield_UserGroup_M
    global excelYield_Month_M
    ansYield_Plant_M = 'ALL'
    ansYield_Machines_M = 'ALL'
    ansYield_Shifts_M   = 'ALL'
    ansYield_UserGroup_M = 'ALL'
    ansYield_StartDate_M = ''
    ansYield_StopDate_M = ''
    ansYield_Month_M = ''
    
    
    if request.method == 'POST':
        ansYield_Plant_M = request.form['Plant']
        ansYield_Machines_M = request.form['Machines']
        ansYield_Shifts_M = request.form['Shifts']
        ansYield_Month_M = request.form['Month12']
        ansYield_UserGroup_M = request.form['UserGroup']
        ansYield_Month_M_Report = ansYield_Month_M
        ansYield_Month_M = ansYield_Month_M.replace("-", "")
     
        print("---------------")
        print(ansYield_Plant_M)
        print(ansYield_Machines_M)
        print(ansYield_Shifts_M)
        print(ansYield_StartDate_M)
        print(ansYield_StopDate_M)
        print(ansYield_UserGroup_M)
        print(ansYield_Month_M)
        
        
        excelYield_Plant_M = ansYield_Plant_M
        excelYield_Machines_M = ansYield_Machines_M
        excelYield_Shifts_M   = ansYield_Shifts_M
        excelYield_Month_M = ansYield_Month_M_Report
       
        excelYield_UserGroup_M = ansYield_UserGroup_M
    
   
    if ansYield_Plant_M == 'ALL':
        ansYield_Plant_M = ''
    else : 
        ansYield_Plant_M = " AND PlantName = '"+ ansYield_Plant_M +"'"
        
    if ansYield_Machines_M == 'ALL':
        ansYield_Machines_M = ''
    else :  
        ansYield_Machines_M = " AND MachineID = '"+ ansYield_Machines_M +"'"
       
    if ansYield_Shifts_M == 'ALL':
        ansYield_Shifts_M = ''
    else : 
        ansYield_Shifts_M = " AND ShiftCode = '"+ ansYield_Shifts_M +"'"
        
    if ansYield_UserGroup_M == 'ALL':
        ansYield_UserGroup_M = ''
    else : 
        ansYield_UserGroup_M = " AND UserGroupID = '"+ ansYield_UserGroup_M +"'"
        
    if ansYield_Month_M == '':
        ansYield_Month_M = ''
    else : 
        ansYield_Month_M = " AND CAST(FORMAT(PostingDate, 'yyyyMM') AS int) = " + ansYield_Month_M 
         
        
    print("---------------")
    print(ansYield_Plant_M)
    print(ansYield_Machines_M)
    print(ansYield_Shifts_M)
    print(ansYield_StartDate_M)
    print(ansYield_StopDate_M)
    print(ansYield_Month_M)
     
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Plant = cnxn.cursor()
    Plant.execute('SELECT PlantName FROM OEE_DB.dbo.Plant  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Machines = cnxn.cursor()
    Machines.execute('SELECT MachineName ,MachineID FROM OEE_DB.dbo.Machines  WHERE DeleteFlag = 1')

    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    Shifts = cnxn.cursor()
    Shifts.execute('SELECT ShiftCodeName , ShiftCodeID FROM OEE_DB.dbo.ShiftCode  WHERE DeleteFlag = 1')
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    UserGroup = cnxn.cursor()
    UserGroup.execute('SELECT UserGroupName , UserGroupID FROM OEE_DB.dbo.UserGroup  WHERE DeleteFlag = 1')

    return render_template('Report_Yield_Montly.html',Plant = Plant , Machines = Machines , Shifts = Shifts,UserGroup=UserGroup,Level=Level,Fname_Lname=Fname_Lname)


@app.route('/Report_OEE_API' ,methods=["GET", "POST"])

def Report_OEE_API():
    global ansOEE_Plant
    global ansOEE_Machines
    global ansOEE_Shifts
    global ansOEE_StartDate
    global ansOEE_StopDate
    global ansOEE_UserGroup
    global where
    
    print("SELECT * FROM OEE_DB.dbo.View_OEEReport  WHERE OEE_Q2 IS NOT NULL " +ansOEE_Plant+ansOEE_Machines+ansOEE_Shifts+ansOEE_StartDate+ansOEE_StopDate)
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportOEE = cnxn.cursor()
    ReportOEE.execute("SELECT * FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL "  +ansOEE_Plant+ansOEE_Machines+ansOEE_Shifts+ansOEE_UserGroup+ansOEE_StartDate+ansOEE_StopDate)
   
    payload = []
    content = {}
    for result in ReportOEE:
        content = {'Plant': result[2], 'work_time': str(result[34] +" / " + result[33] ), 'Posting_Date': str(result[3]),'PD_order': result[4],'Material_number': str(result[5]),'Machine_Text': str(result[8]),'Material_Description': str(result[6]),'MachineID': str(result[7]),'Validate_Speed': str(result[9]),'Q1':  float("{:.2f}".format(result[25])),'Plan_DT': str(result[10]),'Unlan_DT': str(result[11]),'ka_time': str(result[12]),'getwork_time_1': str(result[13]),'getwork_time_2': str(result[14]),'number_of_product': str(result[15]),'number_should_of_product_1':  float("{:.2f}".format(result[16])),'number_should_of_product_2':  float("{:.2f}".format(result[17])),'product_Qty': str(result[18]),'Return_Qty': str(result[19]),'product_Qty_F': str(result[20]),'UserGroup': str(result[32]),'Q2': float("{:.2f}".format(result[38])),'Availability_A1': float("{:.2f}".format(result[21])),'Availability_A2': float("{:.2f}".format(result[22])),'Performance_P1': float("{:.2f}".format(result[23])),'Performance_P2': float("{:.2f}".format(result[24])),'Quality1F': float("{:.2f}".format(result[26])),'Quality2F': float("{:.2f}".format(result[39])),'OEE1': float("{:.2f}".format(result[27])),'OEE2': float("{:.2f}".format(result[28])),'OEE1_F': float("{:.2f}".format(result[29])),'OEE2_F': float("{:.2f}".format(result[30]))}
        payload.append(content)
        content = {}
    #print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201

@app.route('/Report_OEE_API_EXCEL' ,methods=["GET", "POST"])

def Report_OEE_API_EXCEL():
    global ansOEE_Plant
    global ansOEE_Machines
    global ansOEE_Shifts
    global ansOEE_StartDate
    global ansOEE_StopDate
    global ansOEE_UserGroup
    global where
    
    print("SELECT * FROM OEE_DB.dbo.View_OEEReport  WHERE OEE_Q2 IS NOT NULL " +ansOEE_Plant+ansOEE_Machines+ansOEE_Shifts+ansOEE_StartDate+ansOEE_StopDate)
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportOEE = cnxn.cursor()
    ReportOEE.execute("SELECT * FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL "  +ansOEE_Plant+ansOEE_Machines+ansOEE_Shifts+ansOEE_UserGroup+ansOEE_StartDate+ansOEE_StopDate)
   
    payload = []
    content = {}
    for result in ReportOEE:
        #content = {'Plant': result[2], 'work_time': str(result[34] +" / " + result[33] ), 'Posting_Date': str(result[3]),'PD_order': result[4],'Material_number': str(result[5]),'Machine_Text': str(result[8]),'Material_Description': str(result[6]),'MachineID': str(result[7]),'Validate_Speed': str(result[10]),'Q1':  float("{:.2f}".format(result[25])),'Plan_DT': str(result[10]),'Unlan_DT': str(result[11]),'ka_time': str(result[12]),'getwork_time_1': str(result[13]),'getwork_time_2': str(result[14]),'number_of_product': str(result[15]),'number_should_of_product_1':  float("{:.2f}".format(result[16])),'number_should_of_product_2':  float("{:.2f}".format(result[17])),'product_Qty': str(result[18]),'Return_Qty': str(result[19]),'product_Qty_F': str(result[20]),'UserGroup': str(result[32]),'Q2': float("{:.2f}".format(result[38])),'Availability_A1': float("{:.2f}".format(result[21])),'Availability_A2': float("{:.2f}".format(result[22])),'Performance_P1': float("{:.2f}".format(result[23])),'Performance_P2': float("{:.2f}".format(result[24])),'Quality1F': float("{:.2f}".format(result[26])),'Quality2F': float("{:.2f}".format(result[39])),'OEE1': float("{:.2f}".format(result[27])),'OEE2': float("{:.2f}".format(result[28])),'OEE1_F': float("{:.2f}".format(result[29])),'OEE2_F': float("{:.2f}".format(result[30]))}
        content = {'Plant': result[2], 'กะการทำงาน': str(result[34] +" / " + result[33] ), 'Posting Date': str(result[3]),'PD order': result[4],'Material number': str(result[5]),'Machine Text': str(result[8]),'Material Description': str(result[6]),'MachineID': str(result[7]),'Validate Speed': str(result[10]),'Plan DT': str(result[10]),'Unlan DT': str(result[11]),'กะเวลา (min)': result[12],'เวลารับภาระงาน1 (min)': str(result[13]),'เวลารับภาระงาน2 (min)': str(result[14]),'จำนวนชิ้นงานที่ผลิตได้ (Unit)': str(result[15]),'จำนวนชิ้นงานที่ควรจะผลิตได้1 (Unit)':  float("{:.2f}".format(result[16])),'จำนวนชิ้นงานที่ควรจะผลิตได้2 (Unit)':  float("{:.2f}".format(result[17])),'จำนวนชิ้นงานที่ผ่านคุณภาพ (Unit)': str(result[18]),'Return Qty (after 30 day) (Unit)': str(result[19]),'จำนวนชิ้นงานที่ผ่านคุณภาพสุดท้าย (Unit)': str(result[20]),'UserGroup': str(result[32]),'Availability (A1)': float("{:.2f}".format(result[21])),'Availability (A2)': float("{:.2f}".format(result[22])),'Performance (P1)': float("{:.2f}".format(result[23])),'Performance (P2)': float("{:.2f}".format(result[24])),'Quality by PD order (Q1)':  float("{:.2f}".format(result[25])),'Quality by shift (Q2)': float("{:.2f}".format(result[38])),'Quality Final by PD order (Q1)': float("{:.2f}".format(result[26])),'Quality Final by Shift (Q2)': float("{:.2f}".format(result[39])),'OEE1': float("{:.2f}".format(result[27])),'OEE2': float("{:.2f}".format(result[28])),'OEE1_F': float("{:.2f}".format(result[29])),'OEE2_F': float("{:.2f}".format(result[30]))}
        payload.append(content)
        content = {}
    #print(payload)
    return json.dumps(payload, cls = Encoder), 201

    
@app.route('/Report_OEE_Excel')
@flask_login.login_required
def Report_OEE_Excel():
   
    global excelOEE_Plant
    global excelOEE_Machines 
    global excelOEE_Shifts   
    global excelOEE_StartDate 
    global excelOEE_StopDate 
    global excelOEE_UserGroup 
        
    if excelOEE_StopDate == '':
        excelOEE_StopDate = 'ALL'
   
         
    if excelOEE_StartDate == '':
        excelOEE_StartDate = 'ALL'
  
    df = pd.read_json('http://172.30.2.2:5001//Report_OEE_API_EXCEL')
    print(df)
    df.to_excel('OEE_Report1.xlsx',index=False)
    
    time.sleep(1)
    
    wabu = openpyxl.load_workbook('OEE_Report1.xlsx')
    washi = wabu.active
    washi.insert_rows(1,5)
    washi['A1'] = 'Report OEE'
    washi['A1'].alignment = Alignment(vertical='center')
    washi['A1'].alignment = Alignment(horizontal='center')
    washi['A1'].font = Font(color='FFFFFF',
                        size=24,bold=True)
    washi['A1'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi.merge_cells('A1:F1')
    
    washi['A3'] = 'By Plant'
    washi['A3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['A4'] = 'By Machine'
    washi['A4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C3'] = 'By Shifts'
    washi['C3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C4'] = 'By UserGroup'
    washi['C4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E3'] = 'Start Date'
    washi['E3'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E4'] = 'Stop Date'
    washi['E4'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E3'] = 'Start Date'
    washi['E3'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['B3'] = str(excelOEE_Plant)
    washi['B3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['B4'] = str(excelOEE_Machines)
    washi['B4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D3'] = str(excelOEE_Shifts)
    washi['D3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D4'] = str(excelOEE_UserGroup)
    washi['D4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['F3'] = str(excelOEE_StartDate)
    washi['F3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['F3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['F4'] = str(excelOEE_StopDate)
    washi['F4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['F4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi.column_dimensions['A'].width = 16
    washi.column_dimensions['B'].width = 25
    washi.column_dimensions['C'].width = 20
    washi.column_dimensions['D'].width = 20
    washi.column_dimensions['E'].width = 20
    washi.column_dimensions['F'].width = 25
    wabu.save('OEE_Report1.xlsx')
    
    return send_file('..\OEE_Report1.xlsx') 
        
        
@app.route('/Report_Yield_API' ,methods=["GET", "POST"])

def Report_Yield_API():
    global ansYield_Plant
    global ansYield_Machines
    global ansYield_Shifts
    global ansYield_StartDate
    global ansYield_StopDate
    global ansYield_UserGroup
    global where
    
    print("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE Yield IS NOT NULL AND QC_Status = 'P' " +ansYield_Plant+ansYield_Machines+ansYield_Shifts+ansYield_StartDate+ansYield_StopDate)
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportYield = cnxn.cursor()
    ReportYield.execute("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE Yield IS NOT NULL "  +ansYield_Plant+ansYield_Machines+ansYield_Shifts+ansYield_UserGroup+ansYield_StartDate+ansYield_StopDate)
   
    payload = []
    content = {}
    for result in ReportYield:
        content = {'Plant': result[3], 'Posting_Date': str(result[4]),'PD_order': result[5],'Material_number': result[6],'Material_Description': result[7],'MachineName': result[9],'MachineID': result[8],'QA_Status': result[10],'Input_Qty':  result[13],'Output_Qty': result[14],'Return_Qty': result[15],'Shifts': result[11] + '/' + result[12],'Yield': float("{:.2f}".format(result[17])),'Final_Yield': float("{:.2f}".format(result[18]))}
        payload.append(content)
        content = {}
    print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201


@app.route('/Report_Yield_API_EXCEL' ,methods=["GET", "POST"])

def Report_Yield_API_EXCEL():
    global ansYield_Plant
    global ansYield_Machines
    global ansYield_Shifts
    global ansYield_StartDate
    global ansYield_StopDate
    global ansYield_UserGroup
    global where
    
    print("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE Yield IS NOT NULL AND QC_Status = 'P' " +ansYield_Plant+ansYield_Machines+ansYield_Shifts+ansYield_StartDate+ansYield_StopDate)
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportYield = cnxn.cursor()
    ReportYield.execute("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE Yield IS NOT NULL "  +ansYield_Plant+ansYield_Machines+ansYield_Shifts+ansYield_UserGroup+ansYield_StartDate+ansYield_StopDate)
   
    payload = []
    content = {}
    for result in ReportYield:
        content = {'Plant': result[3], 'Posting_Date': str(result[4]),'PD_order': result[5],'Material_number': result[6],'Material_Description': result[7],'MachineName': result[9],'MachineID': result[8],'QA_Status': result[10],'Input_Qty':  result[13],'Output_Qty': result[14],'Return_Qty': result[15],'Shifts': result[11] + '/' + result[12],'Yield': float("{:.2f}".format(result[17])),'Final_Yield': float("{:.2f}".format(result[18]))}
        payload.append(content)
        content = {}
    print(payload)
    return json.dumps(payload, cls = Encoder), 201

@app.route('/Report_Yield_Excel')
@flask_login.login_required
def Report_Yield_Excel():
   
    global excelYield_Plant
    global excelYield_Machines 
    global excelYield_Shifts   
    global excelYield_StartDate 
    global excelYield_StopDate 
    global excelYield_UserGroup 
        
    if excelYield_StopDate == '':
        excelYield_StopDate = 'ALL'
   
         
    if excelYield_StartDate == '':
        excelYield_StartDate = 'ALL'
  
    df = pd.read_json('http://172.30.2.2:5001//Report_Yield_API_EXCEL')
    df.to_excel('Yield_Report.xlsx',index=False)
    
    time.sleep(1)
    wabu = openpyxl.load_workbook('Yield_Report.xlsx')
    washi = wabu.active
    washi.insert_rows(1,5)
    washi['A1'] = 'Report Yield'
    washi['A1'].alignment = Alignment(vertical='center')
    washi['A1'].alignment = Alignment(horizontal='center')
    washi['A1'].font = Font(color='FFFFFF',
                        size=24,bold=True)
    washi['A1'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi.merge_cells('A1:F1')
    
    washi['A3'] = 'By Plant'
    washi['A3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['A4'] = 'By Machine'
    washi['A4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C3'] = 'By Shifts'
    washi['C3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C4'] = 'By UserGroup'
    washi['C4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E3'] = 'Start Date'
    washi['E3'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E4'] = 'Stop Date'
    washi['E4'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E3'] = 'Start Date'
    washi['E3'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['B3'] = str(excelYield_Plant)
    washi['B3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['B4'] = str(excelYield_Machines)
    washi['B4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D3'] = str(excelYield_Shifts)
    washi['D3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D4'] = str(excelYield_UserGroup)
    washi['D4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['F3'] = str(excelYield_StartDate)
    washi['F3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['F3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['F4'] = str(excelYield_StopDate)
    washi['F4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['F4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi.column_dimensions['A'].width = 16
    washi.column_dimensions['B'].width = 25
    washi.column_dimensions['C'].width = 20
    washi.column_dimensions['D'].width = 20
    washi.column_dimensions['E'].width = 20
    washi.column_dimensions['F'].width = 25
    wabu.save('Yield_Report.xlsx')
    
    return send_file('Yield_Report.xlsx') 

@app.route('/Report_M_OEE_API' ,methods=["GET", "POST"])

def Report_M_OEE_API():
    global ansOEE_Plant_M
    global ansOEE_Machines_M
    global ansOEE_Shifts_M
    global ansOEE_StartDate_M
    global ansOEE_StopDate_M
    global ansOEE_UserGroup_M
    global where
    global ansOEE_Month_M
    global ansOEE_Month_M_Report
    
   
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportOEE = cnxn.cursor()
    ReportOEE.execute("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL "  +ansOEE_Plant_M+ansOEE_Machines_M+ansOEE_Shifts_M+ansOEE_UserGroup_M+ansOEE_Month_M)
    print("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL "  +ansOEE_Plant_M+ansOEE_Machines_M+ansOEE_Shifts_M+ansOEE_UserGroup_M+ansOEE_Month_M)
    MachineID = 0 
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    OEEReport_Monthly = cnxn.cursor()
    OEEReport_Monthly.execute('DELETE FROM OEE_DB.dbo.OEEMonthlyReport')
    cnxn.commit()
    Monthly = None
    for i in ReportOEE : 
        allPlanDownTime = 0
        allUnplanDownTime = 0
        allPlannedProductionTime = 0
        allRunTime1 = 0
        allRunTime2 = 0
        allTotalCount = 0
        allIdealCount1 = 0
        allIdealCount2 = 0
        allGoodCount = 0
        allPostReturn = 0
        allFinalGoodCount = 0
        
        Per_PantDownTime = 0
        Per_UnplanDowntime =0
        Per_Downtime = 0
        Per_PantDownTime2 = 0
        Per_UnplanDowntime2 =0
        Per_Downtime2 = 0
        OEE_A1 = 0
        OEE_A2 = 0
        OEE_P1 = 0
        OEE_P2 = 0
        OEE_Q = 0
        OEE_Q_Final = 0
        OEE1Calculation = 0
        OEE2Calculation = 0
        OEE1FinalCalculation = 0
        OEE2FinalCalculation = 0
        
        MachineID = i[0]
        print(MachineID)
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        allMachineID = cnxn.cursor()
        allMachineID.execute("SELECT * FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansOEE_Month_M)
        print("SELECT * FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansOEE_Month_M)
        
        for j in allMachineID:
            print(j[10])
            PlantID = j[1]
            PlantName = j[2]
            MachineID_1 = j[7]
            MachineName = j[8]
            allPlanDownTime += j[10]
            allUnplanDownTime += j[11]
            allPlannedProductionTime  += j[12]
            allRunTime1 += j[13]
            allRunTime2  += j[14]
            allTotalCount  += j[15]
            allIdealCount1  += j[16]
            allIdealCount2  += j[17]
            allGoodCount  += j[18]
            allPostReturn  += j[19]
            allFinalGoodCount  += j[20]
        Per_PantDownTime = allPlanDownTime / allRunTime1
        Per_UnplanDowntime = allUnplanDownTime / allRunTime1
        Per_Downtime = Per_PantDownTime + Per_UnplanDowntime
        Per_PantDownTime2 = allPlanDownTime / allRunTime2
        Per_UnplanDowntime2 = allUnplanDownTime / allRunTime2
        Per_Downtime2 = Per_PantDownTime2 + Per_UnplanDowntime2
        OEE_A1 = (allRunTime2 - allRunTime1)/allRunTime2
        OEE_A2 = (Per_Downtime - allRunTime1) / Per_Downtime
        OEE_P1 = Per_UnplanDowntime2 / Per_Downtime2
        OEE_P2 = Per_UnplanDowntime2 /allTotalCount
        OEE_Q = allGoodCount / allTotalCount
        OEE_Q_Final = allFinalGoodCount / allTotalCount
        OEE1Calculation = OEE_A1 * OEE_P1 * OEE_Q
        OEE2Calculation = OEE_A2 * OEE_P2 * OEE_Q
        OEE1FinalCalculation = OEE_A1 * OEE_P1 * OEE_Q_Final
        OEE2FinalCalculation = OEE_A2 * OEE_P2 * OEE_Q_Final
        
        print("MachineID : " +str(MachineID))
        print("allPlanDownTime : " + str(allPlanDownTime))  
        print("allUnplanDownTime : " + str(allUnplanDownTime))    
        print("allPlannedProductionTime : " + str(allPlannedProductionTime) )   
        print("allRunTime1 : " + str(allRunTime1))    
        print("allRunTime2 : " + str(allRunTime2)  )  
        print("allTotalCount : " + str(allTotalCount)  )  
        print("allIdealCount1 : " + str(allIdealCount1)  )  
        print("allIdealCount2 : " + str(allIdealCount2)  )  
        print("allGoodCount : " + str(allGoodCount)    )
        print("allPostReturn : " + str(allPostReturn)   ) 
        print("allFinalGoodCount : " + str(allFinalGoodCount)   ) 
        
        print("Per_PantDownTime : " +str(Per_PantDownTime))
        print("Per_UnplanDowntime : " +str(Per_UnplanDowntime))
        print("Per_Downtime : " +str(Per_Downtime))
        print("Per_PantDownTime2 : " +str(Per_PantDownTime2))
        print("Per_UnplanDowntime2 : " +str(Per_UnplanDowntime2))
        print("Per_Downtime2 : " +str(Per_Downtime2))
        print("OEE_A1 : " +str(OEE_A1))
        print("OEE_A2 : " +str(OEE_A2))
        print("OEE_P1 : " +str(OEE_P1))
        print("OEE_P2 : " +str(OEE_P2))
        print("OEE_Q : " +str(OEE_Q))
        print("OEE_Q_Final : " +str(OEE_Q_Final))
        print("OEE1Calculation : " +str(OEE1Calculation))
        print("OEE2Calculation : " +str(OEE2Calculation))
        print("OEE1FinalCalculation : " +str(OEE1FinalCalculation))
        print("OEE2FinalCalculation : " +str(OEE2FinalCalculation))

        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        OEEReport_Monthly = cnxn.cursor()
        OEEReport_Monthly.execute('INSERT INTO OEE_DB.dbo.OEEMonthlyReport (Monthly, PlantID, PlantName, MachineID, MachineName, PlanDownTime, UnplanDownTime, RunTime1, RunTime2, Per_PantDownTime, Per_UnplanDowntime, Per_Downtime, Per_PantDownTime2, Per_UnplanDowntime2, Per_Downtime2, TotalCount, IdealCount1, IdealCount2, GoodCount, PostReturn, FinalGoodCount, OEE_A1, OEE_A2, OEE_P1, OEE_P2, OEE_Q, OEE_Q_Final, OEE1Calculation, OEE2Calculation, OEE1FinalCalculation, OEE2FinalCalculation) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,( 'mm-12'  , PlantID , PlantName  , MachineID_1  , MachineName  , allPlanDownTime, allUnplanDownTime, allRunTime1, allRunTime2, Per_PantDownTime, Per_UnplanDowntime, Per_Downtime, Per_PantDownTime2, Per_UnplanDowntime2, Per_Downtime2, allTotalCount, allIdealCount1, allIdealCount2, allGoodCount, allPostReturn, allFinalGoodCount, OEE_A1, OEE_A2, OEE_P1, OEE_P2, OEE_Q, OEE_Q_Final, OEE1Calculation, OEE2Calculation, OEE1FinalCalculation, OEE2FinalCalculation ))
        cnxn.commit()
        
        print("---------------------------------")
        
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    OEEReport_Monthly_Show = cnxn.cursor()
    OEEReport_Monthly_Show.execute('Select * from OEE_DB.dbo.OEEMonthlyReport')    
    payload = []
    content = {}
    if ansOEE_Month_M :
        for result in OEEReport_Monthly_Show:
            
            content = {'Month': ansOEE_Month_M_Report , 'Plant': result[3],'MachineID': str(result[4]),'Machine_Text': str(result[5]),'PlanDownTime':  result[6],'UnplanDownTime': str(result[7]),'RunTime1': str(result[8]),'RunTime2': str(result[9]),'Per_PantDownTime':  float("{:.2f}".format(result[10])),'Per_UnplanDowntime':  float("{:.2f}".format(result[11])),'Per_Downtime': float("{:.2f}".format(result[12])),'Per_PantDownTime2': float("{:.2f}".format(result[13])),'Per_UnplanDowntime2': float("{:.2f}".format(result[14])),'Per_Downtime2': float("{:.2f}".format(result[15])),'TotalCount': float("{:.2f}".format(result[16])),'IdealCount1': float("{:.2f}".format(result[17])),'IdealCount2': float("{:.2f}".format(result[18])),'GoodCount': float("{:.2f}".format(result[19])),'PostReturn': float("{:.2f}".format(result[20])),'FinalGoodCount': float("{:.2f}".format(result[21])),'OEE_A1': float("{:.2f}".format(result[22])),'OEE_A2': float("{:.2f}".format(result[23])),'OEE_P1': float("{:.2f}".format(result[24])),'OEE_P2': float("{:.2f}".format(result[25])),'OEE_Q': float("{:.2f}".format(result[26])),'OEE_Q_Finnal': float("{:.2f}".format(result[27])),'OEE1Calculation': float("{:.2f}".format(result[28])),'OEE2Calculation': float("{:.2f}".format(result[29])),'OEE1FinalCalculation': float("{:.2f}".format(result[30])),'OEE2FinalCalculation': float("{:.2f}".format(result[31]))}
            payload.append(content)
            content = {}
        
    #print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201

@app.route('/Report_M_OEE_API_Excel' ,methods=["GET", "POST"])

def Report_M_OEE_API_Excel():
    global ansOEE_Plant_M
    global ansOEE_Machines_M
    global ansOEE_Shifts_M
    global ansOEE_StartDate_M
    global ansOEE_StopDate_M
    global ansOEE_UserGroup_M
    global where
    global ansOEE_Month_M
    global ansOEE_Month_M_Report
    
   
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportOEE = cnxn.cursor()
    ReportOEE.execute("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL "  +ansOEE_Plant_M+ansOEE_Machines_M+ansOEE_Shifts_M+ansOEE_UserGroup_M+ansOEE_Month_M)
    print("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL "  +ansOEE_Plant_M+ansOEE_Machines_M+ansOEE_Shifts_M+ansOEE_UserGroup_M+ansOEE_Month_M)
    MachineID = 0 
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    OEEReport_Monthly = cnxn.cursor()
    OEEReport_Monthly.execute('DELETE FROM OEE_DB.dbo.OEEMonthlyReport')
    cnxn.commit()
    Monthly = None
    for i in ReportOEE : 
        allPlanDownTime = 0
        allUnplanDownTime = 0
        allPlannedProductionTime = 0
        allRunTime1 = 0
        allRunTime2 = 0
        allTotalCount = 0
        allIdealCount1 = 0
        allIdealCount2 = 0
        allGoodCount = 0
        allPostReturn = 0
        allFinalGoodCount = 0
        
        Per_PantDownTime = 0
        Per_UnplanDowntime =0
        Per_Downtime = 0
        Per_PantDownTime2 = 0
        Per_UnplanDowntime2 =0
        Per_Downtime2 = 0
        OEE_A1 = 0
        OEE_A2 = 0
        OEE_P1 = 0
        OEE_P2 = 0
        OEE_Q = 0
        OEE_Q_Final = 0
        OEE1Calculation = 0
        OEE2Calculation = 0
        OEE1FinalCalculation = 0
        OEE2FinalCalculation = 0
        
        MachineID = i[0]
        print(MachineID)
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        allMachineID = cnxn.cursor()
        allMachineID.execute("SELECT * FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansOEE_Month_M)
        print("SELECT * FROM OEE_DB.dbo.View_OEEReport WHERE OEE_Q2 IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansOEE_Month_M)
        
        for j in allMachineID:
            print(j[10])
            PlantID = j[1]
            PlantName = j[2]
            MachineID_1 = j[7]
            MachineName = j[8]
            allPlanDownTime += j[10]
            allUnplanDownTime += j[11]
            allPlannedProductionTime  += j[12]
            allRunTime1 += j[13]
            allRunTime2  += j[14]
            allTotalCount  += j[15]
            allIdealCount1  += j[16]
            allIdealCount2  += j[17]
            allGoodCount  += j[18]
            allPostReturn  += j[19]
            allFinalGoodCount  += j[20]
        Per_PantDownTime = allPlanDownTime / allRunTime1
        Per_UnplanDowntime = allUnplanDownTime / allRunTime1
        Per_Downtime = Per_PantDownTime + Per_UnplanDowntime
        Per_PantDownTime2 = allPlanDownTime / allRunTime2
        Per_UnplanDowntime2 = allUnplanDownTime / allRunTime2
        Per_Downtime2 = Per_PantDownTime2 + Per_UnplanDowntime2
        OEE_A1 = (allRunTime2 - allRunTime1)/allRunTime2
        OEE_A2 = (Per_Downtime - allRunTime1) / Per_Downtime
        OEE_P1 = Per_UnplanDowntime2 / Per_Downtime2
        OEE_P2 = Per_UnplanDowntime2 /allTotalCount
        OEE_Q = allGoodCount / allTotalCount
        OEE_Q_Final = allFinalGoodCount / allTotalCount
        OEE1Calculation = OEE_A1 * OEE_P1 * OEE_Q
        OEE2Calculation = OEE_A2 * OEE_P2 * OEE_Q
        OEE1FinalCalculation = OEE_A1 * OEE_P1 * OEE_Q_Final
        OEE2FinalCalculation = OEE_A2 * OEE_P2 * OEE_Q_Final
        
        print("MachineID : " +str(MachineID))
        print("allPlanDownTime : " + str(allPlanDownTime))  
        print("allUnplanDownTime : " + str(allUnplanDownTime))    
        print("allPlannedProductionTime : " + str(allPlannedProductionTime) )   
        print("allRunTime1 : " + str(allRunTime1))    
        print("allRunTime2 : " + str(allRunTime2)  )  
        print("allTotalCount : " + str(allTotalCount)  )  
        print("allIdealCount1 : " + str(allIdealCount1)  )  
        print("allIdealCount2 : " + str(allIdealCount2)  )  
        print("allGoodCount : " + str(allGoodCount)    )
        print("allPostReturn : " + str(allPostReturn)   ) 
        print("allFinalGoodCount : " + str(allFinalGoodCount)   ) 
        
        print("Per_PantDownTime : " +str(Per_PantDownTime))
        print("Per_UnplanDowntime : " +str(Per_UnplanDowntime))
        print("Per_Downtime : " +str(Per_Downtime))
        print("Per_PantDownTime2 : " +str(Per_PantDownTime2))
        print("Per_UnplanDowntime2 : " +str(Per_UnplanDowntime2))
        print("Per_Downtime2 : " +str(Per_Downtime2))
        print("OEE_A1 : " +str(OEE_A1))
        print("OEE_A2 : " +str(OEE_A2))
        print("OEE_P1 : " +str(OEE_P1))
        print("OEE_P2 : " +str(OEE_P2))
        print("OEE_Q : " +str(OEE_Q))
        print("OEE_Q_Final : " +str(OEE_Q_Final))
        print("OEE1Calculation : " +str(OEE1Calculation))
        print("OEE2Calculation : " +str(OEE2Calculation))
        print("OEE1FinalCalculation : " +str(OEE1FinalCalculation))
        print("OEE2FinalCalculation : " +str(OEE2FinalCalculation))

        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        OEEReport_Monthly = cnxn.cursor()
        OEEReport_Monthly.execute('INSERT INTO OEE_DB.dbo.OEEMonthlyReport (Monthly, PlantID, PlantName, MachineID, MachineName, PlanDownTime, UnplanDownTime, RunTime1, RunTime2, Per_PantDownTime, Per_UnplanDowntime, Per_Downtime, Per_PantDownTime2, Per_UnplanDowntime2, Per_Downtime2, TotalCount, IdealCount1, IdealCount2, GoodCount, PostReturn, FinalGoodCount, OEE_A1, OEE_A2, OEE_P1, OEE_P2, OEE_Q, OEE_Q_Final, OEE1Calculation, OEE2Calculation, OEE1FinalCalculation, OEE2FinalCalculation) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)' ,(ansOEE_Month_M_Report , PlantID , PlantName  , MachineID_1  , MachineName  , allPlanDownTime, allUnplanDownTime, allRunTime1, allRunTime2, Per_PantDownTime, Per_UnplanDowntime, Per_Downtime, Per_PantDownTime2, Per_UnplanDowntime2, Per_Downtime2, allTotalCount, allIdealCount1, allIdealCount2, allGoodCount, allPostReturn, allFinalGoodCount, OEE_A1, OEE_A2, OEE_P1, OEE_P2, OEE_Q, OEE_Q_Final, OEE1Calculation, OEE2Calculation, OEE1FinalCalculation, OEE2FinalCalculation ))
        cnxn.commit()
        
        print("---------------------------------")
        
        
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    OEEReport_Monthly_Show = cnxn.cursor()
    OEEReport_Monthly_Show.execute('Select * from OEE_DB.dbo.OEEMonthlyReport')    
    payload = []
    content = {}
    if ansOEE_Month_M :
        for result in OEEReport_Monthly_Show:
            
            content = {'Month': ansOEE_Month_M_Report , 'Plant': result[3],'MachineID': str(result[4]),'Machine_Text': str(result[5]),'PlanDownTime':  result[6],'UnplanDownTime': str(result[7]),'RunTime1': str(result[8]),'RunTime2': str(result[9]),'Per_PantDownTime':  float("{:.2f}".format(result[10])),'Per_UnplanDowntime':  float("{:.2f}".format(result[11])),'Per_Downtime': float("{:.2f}".format(result[12])),'Per_PantDownTime2': float("{:.2f}".format(result[13])),'Per_UnplanDowntime2': float("{:.2f}".format(result[14])),'Per_Downtime2': float("{:.2f}".format(result[15])),'TotalCount': float("{:.2f}".format(result[16])),'IdealCount1': float("{:.2f}".format(result[17])),'IdealCount2': float("{:.2f}".format(result[18])),'GoodCount': float("{:.2f}".format(result[19])),'PostReturn': float("{:.2f}".format(result[20])),'FinalGoodCount': float("{:.2f}".format(result[21])),'OEE_A1': float("{:.2f}".format(result[22])),'OEE_A2': float("{:.2f}".format(result[23])),'OEE_P1': float("{:.2f}".format(result[24])),'OEE_P2': float("{:.2f}".format(result[25])),'OEE_Q': float("{:.2f}".format(result[26])),'OEE_Q_Finnal': float("{:.2f}".format(result[27])),'OEE1Calculation': float("{:.2f}".format(result[28])),'OEE2Calculation': float("{:.2f}".format(result[29])),'OEE1FinalCalculation': float("{:.2f}".format(result[30])),'OEE2FinalCalculation': float("{:.2f}".format(result[31]))}
            payload.append(content)
            content = {}
        
    #print(payload)
    return json.dumps(payload, cls = Encoder), 201

@app.route('/Report_OEE_M_Excel')
@flask_login.login_required
def Report_OEE_M_Excel():
   
    global excelOEE_Plant_M
    global excelOEE_Machines_M 
    global excelOEE_Shifts_M  
    global excelOEE_Month_M 
   
  
    df = pd.read_json('http://172.30.2.2:5001/Report_M_OEE_API_Excel')
    df.to_excel('OEE_Montly_Report.xlsx',index=False)
    
    time.sleep(1)
    wabu = openpyxl.load_workbook('OEE_Montly_Report.xlsx')
    washi = wabu.active
    washi.insert_rows(1,5)
    washi['A1'] = 'Report OEE'
    washi['A1'].alignment = Alignment(vertical='center')
    washi['A1'].alignment = Alignment(horizontal='center')
    washi['A1'].font = Font(color='FFFFFF',
                        size=24,bold=True)
    washi['A1'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi.merge_cells('A1:F1')
    
    washi['A3'] = 'By Plant'
    washi['A3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['A4'] = 'By Machine'
    washi['A4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C3'] = 'By Shifts'
    washi['C3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C4'] = 'By UserGroup'
    washi['C4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E3'] = 'Year-Month'
    washi['E3'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    
    washi['B3'] = str(excelOEE_Plant)
    washi['B3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['B4'] = str(excelOEE_Machines)
    washi['B4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D3'] = str(excelOEE_Shifts)
    washi['D3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D4'] = str(excelOEE_UserGroup)
    washi['D4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['F3'] = str(excelOEE_Month_M)
    washi['F3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['F3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
   
    
    washi.column_dimensions['A'].width = 16
    washi.column_dimensions['B'].width = 25
    washi.column_dimensions['C'].width = 20
    washi.column_dimensions['D'].width = 20
    washi.column_dimensions['E'].width = 20
    washi.column_dimensions['F'].width = 25
    wabu.save('OEE_Montly_Report.xlsx')
    
    return send_file('OEE_Montly_Report.xlsx') 

@app.route('/Report_M_Yield_API' ,methods=["GET", "POST"])

def Report_M_Yield_API():
    global ansYield_Plant_M
    global ansYield_Machines_M
    global ansYield_Shifts_M
    global ansYield_StartDate_M
    global ansYield_StopDate_M
    global ansYield_UserGroup_M
    global where
    global ansYield_Month_M
    global ansYield_Month_M_Report
    
   
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportYield = cnxn.cursor()
    ReportYield.execute("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL "  +ansYield_Plant_M+ansYield_Machines_M+ansYield_Shifts_M+ansYield_UserGroup_M+ansYield_Month_M)
    print("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL "  +ansYield_Plant_M+ansYield_Machines_M+ansYield_Shifts_M+ansYield_UserGroup_M+ansYield_Month_M)
    MachineID = 0 
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    YieldReport_Monthly = cnxn.cursor()
    YieldReport_Monthly.execute('DELETE FROM OEE_DB.dbo.YieldMonthlyReport')
    cnxn.commit()
    Monthly = None
    for i in ReportYield : 
        allInputQty = 0
        allOutputQty = 0 
        allReturnQty = 0
    
        MachineID = i[0]
        print(MachineID)
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        allMachineID = cnxn.cursor()
        allMachineID.execute("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansYield_Month_M)
        print("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansYield_Month_M)
        
        for j in allMachineID:
            print(j[10])
            PlantID = j[2]
            PlantName = j[3]
            MachineID_1 = j[8]
            MachineName = j[10]
            allInputQty += j[13]
            allOutputQty += j[14] 
            allReturnQty += j[15]
            
            
       
        Yield = allOutputQty / allInputQty
        FinalYield = (allOutputQty - allReturnQty) / allInputQty
        
        print("allInputQty : " +str(allInputQty))
        print("allOutputQty : " +str(allOutputQty))
        print("allReturnQty : " +str(allReturnQty))
        print("Yield : " +str(Yield))
        print("FinalYield : " + str(FinalYield))  
        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        YieldReport_Monthly = cnxn.cursor()
        YieldReport_Monthly.execute('INSERT INTO OEE_DB.dbo.YieldMonthlyReport (Monthly, PlantID, PlantName, MachineID, MachineName, InputQty, OutputQty, ReturnQty, Yield, FinalYield) VALUES(?,?,?,?,?,?,?,?,?,?)' ,( ansYield_Month_M_Report  , PlantID , PlantName  , MachineID_1  , MachineName  , allInputQty, allOutputQty, allReturnQty, Yield, FinalYield ))
        cnxn.commit()
        
        print("---------------------------------")
        
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    YieldReport_Monthly_Show = cnxn.cursor()
    YieldReport_Monthly_Show.execute('Select * from OEE_DB.dbo.YieldMonthlyReport')    
    payload = []
    content = {}
    if ansYield_Month_M :
        for result in YieldReport_Monthly_Show:
            
            content = {'Month': ansYield_Month_M_Report , 'Plant': result[2],'MachineID': str(result[3]),'Machine_Text': str(result[4]),'InputQty': float("{:.2f}".format(result[5])),'OutputQty': float("{:.2f}".format(result[6])),'ReturnQty': float("{:.2f}".format(result[7])),'Yield': float("{:.2f}".format(result[8])),'FinalYield': float("{:.2f}".format(result[9]))}
            payload.append(content)
            content = {}
        
    #print(payload)
    return json.dumps({"data":payload}, cls = Encoder), 201

@app.route('/Report_M_Yield_API_Excel' ,methods=["GET", "POST"])
def Report_M_Yield_API_Excel():
    global ansYield_Plant_M
    global ansYield_Machines_M
    global ansYield_Shifts_M
    global ansYield_StartDate_M
    global ansYield_StopDate_M
    global ansYield_UserGroup_M
    global where
    global ansYield_Month_M
    global ansYield_Month_M_Report
    
   
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    ReportYield = cnxn.cursor()
    ReportYield.execute("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL "  +ansYield_Plant_M+ansYield_Machines_M+ansYield_Shifts_M+ansYield_UserGroup_M+ansYield_Month_M)
    print("SELECT DISTINCT MachineID FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL "  +ansYield_Plant_M+ansYield_Machines_M+ansYield_Shifts_M+ansYield_UserGroup_M+ansYield_Month_M)
    MachineID = 0 
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    YieldReport_Monthly = cnxn.cursor()
    YieldReport_Monthly.execute('DELETE FROM OEE_DB.dbo.YieldMonthlyReport')
    cnxn.commit()
    Monthly = None
    for i in ReportYield : 
        allInputQty = 0
        allOutputQty = 0 
        allReturnQty = 0
    
        MachineID = i[0]
        print(MachineID)
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        allMachineID = cnxn.cursor()
        allMachineID.execute("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansYield_Month_M)
        print("SELECT * FROM OEE_DB.dbo.View_YieldReport WHERE FinalYield IS NOT NULL AND MachineID ='" + MachineID + "'"  +ansYield_Month_M)
        
        for j in allMachineID:
            print(j[10])
            PlantID = j[2]
            PlantName = j[3]
            MachineID_1 = j[8]
            MachineName = j[10]
            allInputQty += j[13]
            allOutputQty += j[14] 
            allReturnQty += j[15]
            
            
       
        Yield = allOutputQty / allInputQty
        FinalYield = (allOutputQty - allReturnQty) / allInputQty
        
        print("allInputQty : " +str(allInputQty))
        print("allOutputQty : " +str(allOutputQty))
        print("allReturnQty : " +str(allReturnQty))
        print("Yield : " +str(Yield))
        print("FinalYield : " + str(FinalYield))  
        
        cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
        YieldReport_Monthly = cnxn.cursor()
        YieldReport_Monthly.execute('INSERT INTO OEE_DB.dbo.YieldMonthlyReport (Monthly, PlantID, PlantName, MachineID, MachineName, InputQty, OutputQty, ReturnQty, Yield, FinalYield) VALUES(?,?,?,?,?,?,?,?,?,?)' ,( ansYield_Month_M_Report  , PlantID , PlantName  , MachineID_1  , MachineName  , allInputQty, allOutputQty, allReturnQty, Yield, FinalYield ))
        cnxn.commit()
        
        print("---------------------------------")
        
    
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password)
    YieldReport_Monthly_Show = cnxn.cursor()
    YieldReport_Monthly_Show.execute('Select * from OEE_DB.dbo.YieldMonthlyReport')    
    payload = []
    content = {}
    if ansYield_Month_M :
        for result in YieldReport_Monthly_Show:
            
            content = {'Month': ansYield_Month_M_Report , 'Plant': result[2],'MachineID': str(result[3]),'Machine_Text': str(result[4]),'InputQty': float("{:.2f}".format(result[5])),'OutputQty': float("{:.2f}".format(result[6])),'ReturnQty': float("{:.2f}".format(result[7])),'Yield': float("{:.2f}".format(result[8])),'FinalYield': float("{:.2f}".format(result[9]))}
            payload.append(content)
            content = {}
        

    return json.dumps(payload, cls = Encoder), 201

@app.route('/Report_Yield_M_Excel')
@flask_login.login_required
def Report_Yield_M_Excel():
   
    global excelYield_Plant_M
    global excelYield_Machines_M 
    global excelYield_Shifts_M  
    global excelYield_Month_M 
   
  
    df = pd.read_json('http://172.30.2.2:5001/Report_M_Yield_API_Excel')
    df.to_excel('Yield_Montly_Report.xlsx',index=False)

    time.sleep(1)
    wabu = openpyxl.load_workbook('Yield_Montly_Report.xlsx')
    washi = wabu.active
    washi.insert_rows(1,5)
    washi['A1'] = 'Report Yield'
    washi['A1'].alignment = Alignment(vertical='center')
    washi['A1'].alignment = Alignment(horizontal='center')
    washi['A1'].font = Font(color='FFFFFF',
                        size=24,bold=True)
    washi['A1'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi.merge_cells('A1:F1')
    
    washi['A3'] = 'By Plant'
    washi['A3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['A4'] = 'By Machine'
    washi['A4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['A4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C3'] = 'By Shifts'
    washi['C3'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['C4'] = 'By UserGroup'
    washi['C4'].font = Font(color='FFFFFF',
                        size=12,bold=True)
    washi['C4'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    washi['E3'] = 'Year-Month'
    washi['E3'].font = Font(color='FFFFFF',
                       size=12,bold=True)
    washi['E3'].fill = PatternFill(patternType='solid',fgColor='154360')
    
    
    washi['B3'] = str(excelYield_Plant)
    washi['B3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['B4'] = str(excelYield_Machines)
    washi['B4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['B4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D3'] = str(excelYield_Shifts)
    washi['D3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['D4'] = str(excelYield_UserGroup)
    washi['D4'].font = Font(color='000000',
                       size=12,bold=True)
    washi['D4'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
    washi['F3'] = str(excelYield_Month_M)
    washi['F3'].font = Font(color='000000',
                       size=12,bold=True)
    washi['F3'].fill = PatternFill(patternType='solid',fgColor='AED6F1')
    
   
    washi.column_dimensions['A'].width = 16
    washi.column_dimensions['B'].width = 25
    washi.column_dimensions['C'].width = 20
    washi.column_dimensions['D'].width = 20
    washi.column_dimensions['E'].width = 20
    washi.column_dimensions['F'].width = 25
    wabu.save('Yield_Montly_Report.xlsx')
    
    return send_file('Yield_Montly_Report.xlsx') 

#-------------------- main --------------------------------------

if __name__ == "__main__":
    app.run(host='0.0.0.0', debug=True ,port=5001)